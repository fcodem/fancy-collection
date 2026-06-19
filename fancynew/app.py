import os
import sys
import uuid

# Always use the fancynew folder as working directory (safe even if launched elsewhere)
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
if os.getcwd() != _APP_DIR:
    os.chdir(_APP_DIR)
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)
from datetime import date, datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from werkzeug.utils import secure_filename

from models import db, User, Customer, ClothingItem, Rental, RentalItem, Invoice, Payment, Booking, BookingItem, Staff, StaffAttendance, Supplier, SupplierPurchase, CustomCategory, StaffLoginRequest, UserSession

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "cloth-rental-dev-key-change-in-production")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "cloth_rental.db"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


db.init_app(app)


# ─── Auth helpers ──────────────────────────────────────────────────────────────

def get_current_user():
    uid = session.get("user_id")
    sid = session.get("session_id")
    if not uid or not sid:
        return None
    us = UserSession.query.filter_by(session_id=sid, user_id=uid, active=True).first()
    if not us:
        return None
    us.last_seen = datetime.utcnow()
    db.session.commit()
    return User.query.get(uid)


def establish_user_login(user):
    """Create a tracked server-side session after successful login."""
    UserSession.query.filter_by(user_id=user.id, active=True).update(
        {"active": False, "ended_at": datetime.utcnow()},
        synchronize_session=False,
    )
    sid = uuid.uuid4().hex
    us = UserSession(user_id=user.id, session_id=sid, active=True)
    db.session.add(us)
    db.session.commit()
    session["user_id"] = user.id
    session["session_id"] = sid
    session.pop("pending_login_token", None)


def end_user_session(session_row=None, ended_by=None):
    """End one session or the current browser session."""
    if session_row is None:
        sid = session.get("session_id")
        if not sid:
            return
        session_row = UserSession.query.filter_by(session_id=sid, active=True).first()
    if session_row and session_row.active:
        session_row.active = False
        session_row.ended_at = datetime.utcnow()
        if ended_by:
            session_row.ended_by_id = ended_by.id
        db.session.commit()


def get_active_staff_sessions():
    return (
        UserSession.query.filter_by(active=True)
        .join(User, UserSession.user_id == User.id)
        .filter(User.role == "staff")
        .order_by(UserSession.login_at.desc())
        .all()
    )


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not get_current_user():
            session.clear()
            flash("Please log in to continue.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def owner_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        u = get_current_user()
        if not u:
            flash("Please log in to continue.", "error")
            return redirect(url_for("login"))
        if u.role != "owner":
            flash("Access denied. Owner permission required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


LOGIN_REQUEST_TTL_MINUTES = 30


def expire_old_login_requests():
    cutoff = datetime.utcnow() - timedelta(minutes=LOGIN_REQUEST_TTL_MINUTES)
    StaffLoginRequest.query.filter(
        StaffLoginRequest.status == "pending",
        StaffLoginRequest.requested_at < cutoff,
    ).update({"status": "expired"}, synchronize_session=False)
    db.session.commit()


def create_staff_login_request(user):
    expire_old_login_requests()
    StaffLoginRequest.query.filter_by(user_id=user.id, status="pending").update(
        {"status": "expired"}, synchronize_session=False
    )
    req = StaffLoginRequest(user_id=user.id, token=uuid.uuid4().hex)
    db.session.add(req)
    db.session.commit()
    return req


def get_pending_staff_login_requests():
    expire_old_login_requests()
    return (
        StaffLoginRequest.query.filter_by(status="pending")
        .order_by(StaffLoginRequest.requested_at.asc())
        .all()
    )


# ─── Login / Logout ────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if get_current_user():
        return redirect(url_for("dashboard"))
    if session.get("pending_login_token"):
        return redirect(url_for("login_pending"))
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        user = User.query.filter_by(username=username, active=True).first()
        if user and user.check_password(password):
            if user.role == "owner":
                establish_user_login(user)
                flash(f"Welcome back, {user.username}!", "success")
                return redirect(url_for("dashboard"))
            req = create_staff_login_request(user)
            session["pending_login_token"] = req.token
            flash("Login request sent to owner for approval.", "info")
            return redirect(url_for("login_pending"))
        flash("Invalid username or password.", "error")
    return render_template("auth/login.html")


@app.route("/login/pending")
def login_pending():
    if get_current_user():
        return redirect(url_for("dashboard"))
    token = session.get("pending_login_token")
    if not token:
        return redirect(url_for("login"))
    expire_old_login_requests()
    req = StaffLoginRequest.query.filter_by(token=token).first()
    if not req or req.status != "pending":
        session.pop("pending_login_token", None)
        if req and req.status == "rejected":
            flash("Owner denied your login request.", "error")
        elif req and req.status == "expired":
            flash("Login request expired. Please sign in again.", "warning")
        return redirect(url_for("login"))
    return render_template("auth/login_pending.html", username=req.user.username)


@app.route("/api/login-request/status")
def api_login_request_status():
    token = session.get("pending_login_token")
    if not token:
        return jsonify({"status": "none"}), 404
    expire_old_login_requests()
    req = StaffLoginRequest.query.filter_by(token=token).first()
    if not req:
        session.pop("pending_login_token", None)
        return jsonify({"status": "none"})
    if req.status == "approved":
        user = User.query.get(req.user_id)
        if user:
            establish_user_login(user)
        return jsonify({"status": "approved", "redirect": url_for("dashboard")})
    if req.status == "rejected":
        session.pop("pending_login_token", None)
        return jsonify({"status": "rejected", "message": "Owner denied your login request."})
    if req.status == "expired":
        session.pop("pending_login_token", None)
        return jsonify({"status": "expired", "message": "Request expired. Please try again."})
    return jsonify({
        "status": "pending",
        "username": req.user.username,
        "requested_at": req.requested_at.strftime("%d %b %Y, %I:%M %p"),
    })


@app.route("/api/staff-login-requests/pending")
@owner_required
def api_pending_staff_logins():
    requests = get_pending_staff_login_requests()
    return jsonify([
        {
            "id": r.id,
            "username": r.user.username,
            "staff_name": r.user.staff.name if r.user.staff else r.user.username,
            "requested_at": r.requested_at.strftime("%d %b %Y, %I:%M %p"),
            "seconds_ago": int((datetime.utcnow() - r.requested_at).total_seconds()),
        }
        for r in requests
    ])


@app.route("/api/staff-login-request/<int:req_id>/approve", methods=["POST"])
@owner_required
def approve_staff_login(req_id):
    req = StaffLoginRequest.query.get_or_404(req_id)
    if req.status != "pending":
        return jsonify({"error": "Request already handled"}), 400
    req.status = "approved"
    req.resolved_at = datetime.utcnow()
    req.resolved_by_id = get_current_user().id
    db.session.commit()
    return jsonify({"ok": True, "username": req.user.username})


@app.route("/api/staff-login-request/<int:req_id>/reject", methods=["POST"])
@owner_required
def reject_staff_login(req_id):
    req = StaffLoginRequest.query.get_or_404(req_id)
    if req.status != "pending":
        return jsonify({"error": "Request already handled"}), 400
    req.status = "rejected"
    req.resolved_at = datetime.utcnow()
    req.resolved_by_id = get_current_user().id
    db.session.commit()
    return jsonify({"ok": True, "username": req.user.username})


@app.route("/api/staff-sessions/active")
@owner_required
def api_active_staff_sessions():
    rows = get_active_staff_sessions()
    return jsonify([
        {
            "id": s.id,
            "user_id": s.user_id,
            "username": s.user.username,
            "staff_name": s.user.staff.name if s.user.staff else s.user.username,
            "staff_id": s.user.staff_id,
            "login_at": s.login_at.strftime("%d %b %Y, %I:%M %p"),
            "last_seen": s.last_seen.strftime("%d %b %Y, %I:%M %p"),
        }
        for s in rows
    ])


@app.route("/api/staff-session/<int:session_id>/force-logout", methods=["POST"])
@owner_required
def force_logout_staff_session(session_id):
    us = UserSession.query.get_or_404(session_id)
    if not us.active:
        return jsonify({"error": "Session already ended"}), 400
    if us.user.role != "staff":
        return jsonify({"error": "Can only force-logout staff accounts"}), 400
    end_user_session(us, ended_by=get_current_user())
    return jsonify({"ok": True, "username": us.user.username})


@app.route("/api/session/check")
def api_session_check():
    if get_current_user():
        return jsonify({"active": True})
    return jsonify({"active": False}), 401


@app.route("/logout")
def logout():
    end_user_session()
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))

def ensure_owner_exists():
    if not User.query.filter_by(role="owner").first():
        owner = User(username="owner", role="owner")
        owner.set_password("admin123")
        db.session.add(owner)
        db.session.commit()


@app.context_processor
def inject_globals():
    from datetime import datetime, date as d
    current_user = get_current_user()
    mens, womens, jewellery, accessory, other = get_all_categories()
    all_cats = mens + womens + jewellery + accessory + other
    # Low-stock: items with status = available, count available per category; flag if < 3 in any
    try:
        available_count = ClothingItem.query.filter_by(status="available").count()
    except Exception:
        available_count = 0
    # Count overdue deliveries for sidebar badge
    try:
        overdue_delivery_count = Booking.query.filter(
            Booking.delivery_date < d.today(),
            Booking.status == "booked"
        ).count()
    except Exception:
        overdue_delivery_count = 0
    return {
        "now": datetime.utcnow(),
        "today": d.today(),
        "today_iso": d.today().isoformat(),
        "mens_categories": mens,
        "womens_categories": womens,
        "jewellery_categories": jewellery,
        "accessory_categories": accessory,
        "sub_categories": SUB_CATEGORIES,
        "all_sizes": SIZES,
        "all_categories": all_cats,
        "current_user": current_user,
        "is_owner": current_user and current_user.role == "owner",
        "available_inventory_count": available_count,
        "overdue_delivery_count": overdue_delivery_count,
    }

_BASE_MENS = ["Sherwani", "Indowestern", "Jodhpuri", "Coat Suit", "Suit", "Blazer", "Kurta"]
_BASE_WOMENS = ["Saree", "Lehenga", "Gown"]
_BASE_JEWELLERY = ["Jewellery", "Necklace", "Bangles", "Earrings", "Maang Tikka", "Haath Phool", "Anklet", "Nose Ring", "Matha Patti"]
_BASE_ACCESSORY = ["Accessory", "Dupatta", "Belt", "Clutch", "Crown/Tiara"]
SIZES = [str(n) for n in range(32, 59, 2)] + ["Free Size", "Custom"]
SUB_CATEGORIES = ["Premium", "Normal", "Cheap"]
PAYMENT_METHODS = ["cash", "card", "upi", "bank"]


def dress_name_words(q):
    return [w.strip().lower() for w in (q or "").split() if w.strip()]


def dress_name_matches(text, q):
    """True if every word in query appears in text (any order)."""
    text_l = (text or "").lower()
    words = dress_name_words(q)
    if not words:
        return True
    return all(w in text_l for w in words)


def dress_name_sql_filter(q):
    """SQLAlchemy filter: all query words must match name/sku/notes."""
    words = dress_name_words(q)
    if not words:
        return None
    parts = []
    for word in words:
        like = f"%{word}%"
        parts.append(db.or_(
            ClothingItem.name.ilike(like),
            ClothingItem.sku.ilike(like),
            ClothingItem.condition_notes.ilike(like),
        ))
    return db.and_(*parts)


def is_sherwani_category(category):
    return (category or "").strip().lower() == "sherwani"


def dress_display_name(name, category=None, size=None):
    """Sherwani bookings always show booked size in labels."""
    name = (name or "").strip()
    cat = (category or "").strip()
    sz = (size or "").strip()
    if is_sherwani_category(cat) and sz:
        low = name.lower()
        if f"size {sz.lower()}" not in low and f"({sz})" not in name and "· size" not in low:
            return f"{name} · Size {sz}"
    return name


def booking_item_size(bi):
    """Booked size for display; falls back to inventory item size."""
    sz = (bi.size or "").strip()
    if sz:
        return sz
    if bi.item_id:
        item = ClothingItem.query.get(bi.item_id)
        return (item.size or "").strip() if item else ""
    return ""


def booking_items_for_api(booking):
    """Serialize booking items with display_name (Sherwani size included)."""
    items_list = []
    if booking.booking_items:
        for bi in booking.booking_items:
            sz = booking_item_size(bi)
            items_list.append({
                "name": bi.dress_name,
                "display_name": dress_display_name(bi.dress_name, bi.category, sz),
                "category": bi.category or "",
                "size": sz,
            })
    elif booking.dress_name:
        item_obj = ClothingItem.query.get(booking.item_id) if booking.item_id else None
        cat = item_obj.category if item_obj else ""
        sz = item_obj.size if item_obj else ""
        items_list.append({
            "name": booking.dress_name,
            "display_name": dress_display_name(booking.dress_name, cat, sz),
            "category": cat,
            "size": sz or "",
        })
    return items_list


def booking_search_filter(query_text):
    """Match customer/serial/phone OR dress name with any word order."""
    if not query_text:
        return None
    q = query_text.strip()
    like_full = f"%{q}%"
    identity_match = db.or_(
        Booking.customer_name.ilike(like_full),
        Booking.contact_1.ilike(like_full),
        Booking.whatsapp_no.ilike(like_full),
        Booking.monthly_serial.cast(db.String).ilike(like_full),
        Booking.booking_number.ilike(like_full),
    )
    words = dress_name_words(q)
    if not words:
        return identity_match
    dress_parts = []
    for word in words:
        wlike = f"%{word}%"
        dress_parts.append(db.or_(
            Booking.dress_name.ilike(wlike),
            Booking.id.in_(
                db.session.query(BookingItem.booking_id).filter(
                    BookingItem.dress_name.ilike(wlike)
                )
            ),
        ))
    return db.or_(identity_match, db.and_(*dress_parts))


@app.template_filter("dress_label")
def dress_label_filter(name, category="", size=""):
    return dress_display_name(name, category, size)


def get_all_categories():
    """Return category lists merged with any custom categories from DB."""
    try:
        custom = CustomCategory.query.filter_by(active=True).all()
    except Exception:
        custom = []
    mens = list(_BASE_MENS)
    womens = list(_BASE_WOMENS)
    jewellery = list(_BASE_JEWELLERY)
    accessory = list(_BASE_ACCESSORY)
    other = ["Other"]
    for c in custom:
        if c.group == "mens" and c.name not in mens:
            mens.append(c.name)
        elif c.group == "womens" and c.name not in womens:
            womens.append(c.name)
        elif c.group == "jewellery" and c.name not in jewellery:
            jewellery.append(c.name)
        elif c.group == "accessory" and c.name not in accessory:
            accessory.append(c.name)
        elif c.group == "other" and c.name not in other:
            other.append(c.name)
    return mens, womens, jewellery, accessory, other


# Static fallbacks used before app context is available
MENS_CATEGORIES = _BASE_MENS
WOMENS_CATEGORIES = _BASE_WOMENS
JEWELLERY_CATEGORIES = _BASE_JEWELLERY
ACCESSORY_CATEGORIES = _BASE_ACCESSORY
CATEGORIES = MENS_CATEGORIES + WOMENS_CATEGORIES + JEWELLERY_CATEGORIES + ACCESSORY_CATEGORIES + ["Other"]


def generate_number(prefix, model, field):
    today = date.today().strftime("%Y%m%d")
    pattern = f"{prefix}-{today}-%"
    last = model.query.filter(getattr(model, field).like(pattern)).order_by(getattr(model, field).desc()).first()
    if last:
        last_num = int(getattr(last, field).split("-")[-1])
        count = last_num + 1
    else:
        count = 1
    return f"{prefix}-{today}-{count:03d}"


def is_unlucky_serial(n):
    """Return True if the digit sum of n is 4 or 8 (skipped serials)."""
    return sum(int(d) for d in str(n)) in (4, 8)


def next_valid_serial(start):
    """Return the smallest integer >= start whose digit sum is not 4 or 8."""
    n = start
    while is_unlucky_serial(n):
        n += 1
    return n


def serial_position_to_value(pos):
    """Given that we want the pos-th valid serial (1-indexed), return its actual number."""
    count = 0
    n = 1
    while True:
        if not is_unlucky_serial(n):
            count += 1
            if count == pos:
                return n
        n += 1


def generate_item_sku():
    last = db.session.query(db.func.max(ClothingItem.id)).scalar() or 0
    return f"ITM-{last + 1:04d}"


def seed_database():
    if Customer.query.first():
        return

    customers = [
        Customer(name="Priya Sharma", phone="9876543210", email="priya@email.com", address="Mumbai"),
        Customer(name="Rahul Mehta", phone="9123456780", email="rahul@email.com", address="Delhi"),
        Customer(name="Anita Desai", phone="9988776655", email="anita@email.com", address="Pune"),
    ]
    db.session.add_all(customers)

    items = [
        ClothingItem(name="Red Bridal Lehenga", sku="LRG-001", category="Lehenga", item_type="clothing", size="M", color="Red", daily_rate=2500, deposit=10000),
        ClothingItem(name="Royal Blue Sherwani", sku="SHR-001", category="Sherwani", item_type="clothing", size="L", color="Blue", daily_rate=1800, deposit=8000),
        ClothingItem(name="Silk Wedding Saree", sku="SAR-001", category="Saree", item_type="clothing", size="Free Size", color="Gold", daily_rate=1200, deposit=5000),
        ClothingItem(name="Black Tuxedo Suit", sku="SUT-001", category="Suit", item_type="clothing", size="L", color="Black", daily_rate=1500, deposit=6000),
        ClothingItem(name="Evening Gown", sku="GWN-001", category="Gown", item_type="clothing", size="S", color="Navy", daily_rate=2000, deposit=7000),
        ClothingItem(name="Designer Kurta Set", sku="KRT-001", category="Kurta", item_type="clothing", size="M", color="Cream", daily_rate=800, deposit=3000),
        ClothingItem(name="Velvet Blazer", sku="BLZ-001", category="Blazer", item_type="clothing", size="L", color="Maroon", daily_rate=600, deposit=2500),
        ClothingItem(name="Pearl Necklace Bridal Set", sku="JWL-001", category="Necklace", item_type="jewellery", size="Free Size", color="White/Gold", daily_rate=500, deposit=3000),
        ClothingItem(name="Kundan Maang Tikka", sku="JWL-002", category="Maang Tikka", item_type="jewellery", size="Free Size", color="Gold/Red", daily_rate=300, deposit=1500),
        ClothingItem(name="Gold Bangles Set (12pc)", sku="JWL-003", category="Bangles", item_type="jewellery", size="2.6", color="Gold", daily_rate=400, deposit=2000),
        ClothingItem(name="Diamond Choker Set", sku="JWL-004", category="Jewellery", item_type="jewellery", size="Free Size", color="Silver/White", daily_rate=800, deposit=5000),
    ]
    db.session.add_all(items)
    db.session.commit()


# ── Dashboard ──────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    today = date.today()
    total_items = ClothingItem.query.count()
    available_items = ClothingItem.query.filter_by(status="available").count()
    rented_items = ClothingItem.query.filter_by(status="rented").count()
    total_customers = Customer.query.count()
    active_rentals = Rental.query.filter(Rental.status.in_(["active", "overdue"])).count()
    overdue_rentals = Rental.query.filter(
        Rental.status == "active", Rental.end_date < today
    ).count()

    month_start = today.replace(day=1)
    monthly_revenue = db.session.query(db.func.sum(Payment.amount)).filter(
        Payment.paid_at >= datetime.combine(month_start, datetime.min.time())
    ).scalar() or 0

    outstanding = db.session.query(db.func.sum(Invoice.total - Invoice.amount_paid)).filter(
        Invoice.status.in_(["unpaid", "partial"])
    ).scalar() or 0

    recent_rentals = Rental.query.order_by(Rental.created_at.desc()).limit(5).all()
    upcoming_returns = Rental.query.filter(
        Rental.status == "active",
        Rental.end_date >= today,
        Rental.end_date <= today + timedelta(days=3),
    ).order_by(Rental.end_date).limit(5).all()

    overdue_list = Rental.query.filter(
        Rental.status == "active", Rental.end_date < today
    ).order_by(Rental.end_date).limit(5).all()

    inventory_items = ClothingItem.query.order_by(ClothingItem.category, ClothingItem.name).all()

    # Today's booking stats
    today_total_orders = Booking.query.filter(
        Booking.delivery_date == today
    ).count()
    today_delivered = Booking.query.filter(
        Booking.delivery_date == today,
        Booking.status == "delivered"
    ).count()
    today_remaining_delivery = Booking.query.filter(
        Booking.delivery_date == today,
        Booking.status == "booked"
    ).count()
    today_returning = Booking.query.filter(
        Booking.return_date == today,
        Booking.status.in_(["booked", "delivered"])
    ).count()

    # ALL undelivered: today + every overdue previous date
    all_undelivered_count = Booking.query.filter(
        Booking.delivery_date <= today,
        Booking.status == "booked"
    ).count()
    all_undelivered_list = Booking.query.filter(
        Booking.delivery_date <= today,
        Booking.status == "booked"
    ).order_by(Booking.delivery_date.asc(), Booking.delivery_time).all()

    # Lists for today
    today_deliveries_list = Booking.query.filter(
        Booking.delivery_date == today
    ).order_by(Booking.delivery_time).all()
    today_returns_list = Booking.query.filter(
        Booking.return_date == today,
        Booking.status.in_(["booked", "delivered"])
    ).order_by(Booking.return_time).all()

    # Late returns count
    late_return_count = Booking.query.filter(
        Booking.return_date < today,
        Booking.status == "delivered",
    ).count()

    pending_staff_logins = get_pending_staff_login_requests() if get_current_user() and get_current_user().role == "owner" else []
    active_staff_sessions = get_active_staff_sessions() if get_current_user() and get_current_user().role == "owner" else []

    return render_template(
        "dashboard.html",
        stats={
            "total_items": total_items,
            "available_items": available_items,
            "rented_items": rented_items,
            "total_customers": total_customers,
            "active_rentals": active_rentals,
            "overdue_rentals": overdue_rentals,
            "monthly_revenue": monthly_revenue,
            "outstanding": outstanding,
        },
        today_stats={
            "total_orders": today_total_orders,
            "delivered": today_delivered,
            "remaining_delivery": today_remaining_delivery,
            "returning": today_returning,
            "all_undelivered": all_undelivered_count,
        },
        today_deliveries_list=today_deliveries_list,
        today_returns_list=today_returns_list,
        all_undelivered_list=all_undelivered_list,
        late_return_count=late_return_count,
        recent_rentals=recent_rentals,
        upcoming_returns=upcoming_returns,
        overdue_list=overdue_list,
        inventory_items=inventory_items,
        categories=CATEGORIES,
        today=today,
        pending_staff_logins=pending_staff_logins,
        active_staff_sessions=active_staff_sessions,
    )


# ── Inventory ──────────────────────────────────────────────────────────────

@app.route("/inventory")
@login_required
def inventory_list():
    category = request.args.get("category", "")
    status = request.args.get("status", "")
    search = request.args.get("q", "")
    sub_category = request.args.get("sub_category", "")

    query = ClothingItem.query
    if category:
        query = query.filter_by(category=category)
    if status:
        query = query.filter_by(status=status)
    if sub_category:
        query = query.filter_by(sub_category=sub_category)
    if search:
        query = query.filter(
            db.or_(
                ClothingItem.name.ilike(f"%{search}%"),
                ClothingItem.sku.ilike(f"%{search}%"),
            )
        )
    items = query.order_by(ClothingItem.name).all()
    return render_template(
        "inventory/list.html",
        items=items,
        categories=CATEGORIES,
        current_category=category,
        current_status=status,
        search=search,
    )


@app.route("/inventory/search")
@login_required
def inventory_search_page():
    return render_template("inventory/dress_search.html")


@app.route("/api/inventory/search")
@login_required
def api_inventory_search():
    """Live search inventory by name, SKU, notes. Category-first with fallback to other categories."""
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    if not q or len(q) < 1:
        return jsonify({"category_results": [], "other_results": [], "used_fallback": False, "category": category})

    name_filter = dress_name_sql_filter(q)

    def _serialize(items):
        return [{
            "id": i.id,
            "name": i.name,
            "display_name": dress_display_name(i.name, i.category, i.size),
            "sku": i.sku,
            "category": i.category,
            "status": i.status,
            "size": i.size or "",
            "sub_category": i.sub_category or "",
            "photo": i.photo or "",
        } for i in items]

    category_results = []
    other_results = []
    used_fallback = False

    if category:
        q_cat = ClothingItem.query
        if name_filter is not None:
            q_cat = q_cat.filter(name_filter)
        category_results = q_cat.filter(
            ClothingItem.category == category,
        ).order_by(ClothingItem.name).limit(20).all()

        if not category_results:
            used_fallback = True
            q_other = ClothingItem.query
            if name_filter is not None:
                q_other = q_other.filter(name_filter)
            other_results = q_other.filter(
                ClothingItem.category != category,
            ).order_by(ClothingItem.name).limit(20).all()
    else:
        q_all = ClothingItem.query
        if name_filter is not None:
            q_all = q_all.filter(name_filter)
        category_results = q_all.order_by(ClothingItem.name).limit(20).all()

    return jsonify({
        "category": category,
        "category_results": _serialize(category_results),
        "other_results": _serialize(other_results),
        "used_fallback": used_fallback,
    })


@app.route("/api/dress-name/suggest")
@login_required
def api_dress_name_suggest():
    """Autocomplete dress names — words can be entered in any order."""
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    try:
        limit = min(int(request.args.get("limit", 12)), 25)
    except ValueError:
        limit = 12
    if not q:
        return jsonify([])

    query = ClothingItem.query
    name_filter = dress_name_sql_filter(q)
    if name_filter is not None:
        query = query.filter(name_filter)
    if category:
        query = query.filter(ClothingItem.category == category)

    items = query.order_by(ClothingItem.name, ClothingItem.size).limit(limit).all()
    return jsonify([
        {
            "id": i.id,
            "name": i.name,
            "display_name": dress_display_name(i.name, i.category, i.size),
            "category": i.category,
            "size": i.size or "",
            "sku": i.sku,
            "status": i.status,
        }
        for i in items
    ])


@app.route("/api/inventory/photo-search", methods=["POST"])
@login_required
def api_inventory_photo_search():
    """Find inventory items by uploaded photo. Category-first with fallback to other categories."""
    if "photo" not in request.files:
        return jsonify({"error": "No photo uploaded"}), 400
    file = request.files["photo"]
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Invalid file"}), 400
    category = request.form.get("category", "").strip()
    try:
        from PIL import Image
        import io
        img_bytes = file.read()
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB").resize((16, 16))
        pixels = list(img.getdata())
        avg = sum(sum(p) for p in pixels) / (len(pixels) * 3)
        query_hash = sum(
            1 << i for i, p in enumerate(pixels)
            if (p[0] + p[1] + p[2]) / 3 >= avg
        )

        all_items = ClothingItem.query.filter(ClothingItem.photo != None, ClothingItem.photo != "").all()
        scored = []
        for item in all_items:
            photo_path = os.path.join(UPLOAD_FOLDER, item.photo)
            if not os.path.exists(photo_path):
                continue
            try:
                stored_img = Image.open(photo_path).convert("RGB").resize((16, 16))
                s_pixels = list(stored_img.getdata())
                s_avg = sum(sum(p) for p in s_pixels) / (len(s_pixels) * 3)
                stored_hash = sum(
                    1 << i for i, p in enumerate(s_pixels)
                    if (p[0] + p[1] + p[2]) / 3 >= s_avg
                )
                diff = bin(query_hash ^ stored_hash).count("1")
                similarity = round((256 - diff) / 256 * 100, 1)
                if similarity >= 40:
                    scored.append((similarity, item))
            except Exception:
                continue

        scored.sort(key=lambda x: -x[0])

        def _item_dict(sim, item):
            return {
                "id": item.id,
                "name": item.name,
                "sku": item.sku,
                "category": item.category,
                "status": item.status,
                "size": item.size or "",
                "photo": item.photo or "",
                "similarity": sim,
            }

        category_results = []
        other_results = []
        used_fallback = False

        if category:
            cat_scored = [(s, i) for s, i in scored if i.category == category][:10]
            if cat_scored:
                category_results = [_item_dict(s, i) for s, i in cat_scored]
            else:
                used_fallback = True
                other_scored = [(s, i) for s, i in scored if i.category != category][:10]
                other_results = [_item_dict(s, i) for s, i in other_scored]
        else:
            category_results = [_item_dict(s, i) for s, i in scored[:10]]

        return jsonify({
            "ok": True,
            "category": category,
            "category_results": category_results,
            "other_results": other_results,
            "used_fallback": used_fallback,
            "results": category_results + other_results,
        })
    except ImportError:
        return jsonify({"error": "PIL not installed. Install with: pip install Pillow"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/inventory/add", methods=["GET", "POST"])
@owner_required
def inventory_add():
    if request.method == "POST":
        photo_filename = ""
        if "photo" in request.files:
            file = request.files["photo"]
            if file and file.filename and allowed_file(file.filename):
                ext = file.filename.rsplit(".", 1)[1].lower()
                photo_filename = f"{uuid.uuid4().hex}.{ext}"
                file.save(os.path.join(app.config["UPLOAD_FOLDER"], photo_filename))

        category = request.form["category"]
        item_type = "clothing"
        jewellery_cats = ["Jewellery", "Necklace", "Bangles", "Earrings", "Maang Tikka",
                          "Haath Phool", "Anklet", "Nose Ring", "Matha Patti"]
        accessory_cats = ["Accessory", "Dupatta", "Belt", "Clutch", "Crown/Tiara"]
        if category in jewellery_cats:
            item_type = "jewellery"
        elif category in accessory_cats:
            item_type = "accessory"

        sub_category = request.form.get("sub_category", "Normal")

        # For men's clothing: each size is a separate product
        if category in MENS_CATEGORIES:
            sizes_list = request.form.getlist("sizes[]")
            if not sizes_list:
                flash("Please select at least one size for men's clothing.", "error")
                return redirect(url_for("inventory_add"))
            name = request.form["name"].strip()
            count = 0
            for sz in sizes_list:
                item = ClothingItem(
                    name=name,
                    sku=generate_item_sku(),
                    category=category,
                    size=sz,
                    color="",
                    daily_rate=float(request.form.get("daily_rate", 0)),
                    deposit=float(request.form.get("deposit", 0)),
                    condition_notes=request.form.get("condition_notes", ""),
                    item_type=item_type,
                    photo=photo_filename,
                    sub_category=sub_category,
                )
                db.session.add(item)
                count += 1
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                flash(f"Error adding item: {e}", "error")
                return redirect(url_for("inventory_add"))
            flash(f"'{name}' added with {count} size(s) as separate products.", "success")
            return redirect(url_for("inventory_list"))
        else:
            size_val = request.form.get("size", "")
            item = ClothingItem(
                name=request.form["name"].strip(),
                sku=generate_item_sku(),
                category=category,
                size=size_val,
                color=request.form.get("color", "") if category in WOMENS_CATEGORIES else "",
                daily_rate=float(request.form.get("daily_rate", 0)),
                deposit=float(request.form.get("deposit", 0)),
                condition_notes=request.form.get("condition_notes", ""),
                item_type=item_type,
                photo=photo_filename,
                sub_category=sub_category,
            )
            db.session.add(item)
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                flash(f"Error adding item: {e}", "error")
                return redirect(url_for("inventory_add"))
            flash(f"Item '{item.name}' added successfully.", "success")
            return redirect(url_for("inventory_list"))

    return render_template("inventory/form.html", item=None, categories=CATEGORIES, sizes=SIZES)


@app.route("/inventory/<int:id>/edit", methods=["GET", "POST"])
@owner_required
def inventory_edit(id):
    item = ClothingItem.query.get_or_404(id)
    if request.method == "POST":
        item.name = request.form["name"].strip()
        item.category = request.form["category"]
        sizes_list = request.form.getlist("sizes[]")
        item.size = ", ".join(sizes_list) if sizes_list else request.form.get("size", "")
        item.color = request.form.get("color", "") if item.category in WOMENS_CATEGORIES else ""
        item.daily_rate = float(request.form.get("daily_rate", 0))
        item.deposit = float(request.form.get("deposit", 0))
        item.status = request.form.get("status", item.status)
        item.condition_notes = request.form.get("condition_notes", "")

        category = item.category
        jewellery_cats = ["Jewellery", "Necklace", "Bangles", "Earrings", "Maang Tikka",
                          "Haath Phool", "Anklet", "Nose Ring", "Matha Patti"]
        accessory_cats = ["Accessory", "Dupatta", "Belt", "Clutch", "Crown/Tiara"]
        if category in jewellery_cats:
            item.item_type = "jewellery"
        elif category in accessory_cats:
            item.item_type = "accessory"
        else:
            item.item_type = "clothing"

        if "photo" in request.files:
            file = request.files["photo"]
            if file and file.filename and allowed_file(file.filename):
                if item.photo:
                    old_path = os.path.join(app.config["UPLOAD_FOLDER"], item.photo)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                ext = file.filename.rsplit(".", 1)[1].lower()
                photo_filename = f"{uuid.uuid4().hex}.{ext}"
                file.save(os.path.join(app.config["UPLOAD_FOLDER"], photo_filename))
                item.photo = photo_filename

        db.session.commit()
        flash("Item updated successfully.", "success")
        return redirect(url_for("inventory_list"))

    return render_template("inventory/form.html", item=item, categories=CATEGORIES, sizes=SIZES)


@app.route("/inventory/<int:id>/delete", methods=["POST"])
@owner_required
def inventory_delete(id):
    item = ClothingItem.query.get_or_404(id)
    if item.status == "rented":
        flash("Cannot delete an item that is currently rented.", "error")
        return redirect(url_for("inventory_list"))
    db.session.delete(item)
    db.session.commit()
    flash("Item deleted.", "success")
    return redirect(url_for("inventory_list"))


# ── Customers ──────────────────────────────────────────────────────────────

@app.route("/customers")
@login_required
def customers_list():
    search = request.args.get("q", "")
    category_filter = request.args.get("category", "")
    query = Customer.query
    if search:
        query = query.filter(
            db.or_(
                Customer.name.ilike(f"%{search}%"),
                Customer.phone.ilike(f"%{search}%"),
            )
        )
    customers = query.order_by(Customer.name).all()

    # Filter by booking category if specified
    if category_filter:
        customer_ids = set()
        bookings = Booking.query.filter(Booking.status != "cancelled").all()
        for b in bookings:
            if b.booking_items:
                for bi in b.booking_items:
                    if bi.category == category_filter:
                        # Match by customer_name since bookings don't link to Customer model
                        customer_ids.add(b.customer_name.lower())
            elif b.item_id:
                item = ClothingItem.query.get(b.item_id)
                if item and item.category == category_filter:
                    customer_ids.add(b.customer_name.lower())
        customers = [c for c in customers if c.name.lower() in customer_ids]

    return render_template("customers/list.html", customers=customers, search=search, categories=CATEGORIES, current_category=category_filter)


@app.route("/customers/add", methods=["GET", "POST"])
@login_required
def customers_add():
    if request.method == "POST":
        customer = Customer(
            name=request.form["name"].strip(),
            phone=request.form["phone"].strip(),
            email=request.form.get("email", "").strip(),
            address=request.form.get("address", "").strip(),
            id_proof=request.form.get("id_proof", "").strip(),
            notes=request.form.get("notes", "").strip(),
        )
        db.session.add(customer)
        db.session.commit()
        flash(f"Customer '{customer.name}' added.", "success")
        return redirect(url_for("customers_list"))

    return render_template("customers/form.html", customer=None)


@app.route("/customers/<int:id>")
@login_required
def customers_view(id):
    customer = Customer.query.get_or_404(id)
    rentals = sorted(customer.rentals, key=lambda r: r.created_at or datetime.min, reverse=True)
    return render_template("customers/view.html", customer=customer, rentals=rentals)


@app.route("/customers/<int:id>/edit", methods=["GET", "POST"])
@login_required
def customers_edit(id):
    customer = Customer.query.get_or_404(id)
    if request.method == "POST":
        customer.name = request.form["name"].strip()
        customer.phone = request.form["phone"].strip()
        customer.email = request.form.get("email", "").strip()
        customer.address = request.form.get("address", "").strip()
        customer.id_proof = request.form.get("id_proof", "").strip()
        customer.notes = request.form.get("notes", "").strip()
        db.session.commit()
        flash("Customer updated.", "success")
        return redirect(url_for("customers_view", id=id))

    return render_template("customers/form.html", customer=customer)


@app.route("/customers/<int:id>/delete", methods=["POST"])
@login_required
def customers_delete(id):
    customer = Customer.query.get_or_404(id)
    if len(customer.rentals) > 0:  # customer.rentals is a list backref, not a query
        flash("Cannot delete customer with rental history.", "error")
        return redirect(url_for("customers_view", id=id))
    db.session.delete(customer)
    db.session.commit()
    flash("Customer deleted.", "success")
    return redirect(url_for("customers_list"))


@app.route("/customers/export/whatsapp")
@login_required
def export_customers_whatsapp():
    """Export customer contacts as Excel formatted for WhatsApp bulk API tools."""
    category = request.args.get("category", "").strip()
    from flask import Response
    import io

    # Group contacts by category for multi-sheet export
    if category:
        sheets = {category: _get_customers_by_category(category)}
    else:
        sheets = {}
        bookings = Booking.query.filter(Booking.status != "cancelled").all()
        all_cats = set()
        for b in bookings:
            if b.booking_items:
                for bi in b.booking_items:
                    if bi.category:
                        all_cats.add(bi.category)
            elif b.item_id:
                item = ClothingItem.query.get(b.item_id)
                if item and item.category:
                    all_cats.add(item.category)
        if not all_cats:
            all_cats = {"All Customers"}
            sheets["All Customers"] = _get_customers_by_category(None)
        else:
            for cat in sorted(all_cats):
                contacts = _get_customers_by_category(cat)
                if contacts:
                    sheets[cat] = contacts

    headers = ["Name", "Phone", "Country Code", "WhatsApp Number", "Category", "Address"]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, contacts in sheets.items():
            safe_name = sheet_name[:31].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=safe_name)
            # Header row — WhatsApp API standard columns
            header_fill = PatternFill("solid", fgColor="7B1F45")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for row_idx, c in enumerate(contacts, 2):
                phone_digits = "".join(ch for ch in str(c["phone"]) if ch.isdigit())
                if len(phone_digits) > 10:
                    phone_digits = phone_digits[-10:]
                ws.cell(row=row_idx, column=1, value=c["name"])
                ws.cell(row=row_idx, column=2, value=phone_digits)
                ws.cell(row=row_idx, column=3, value="91")
                ws.cell(row=row_idx, column=4, value=c["whatsapp_number"])
                ws.cell(row=row_idx, column=5, value=sheet_name if category else ", ".join(c.get("categories", [])) or "General")
                ws.cell(row=row_idx, column=6, value=c.get("address", ""))
            for col in range(1, 7):
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col)].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"whatsapp_contacts_{category or 'all'}_{date.today().isoformat()}.xlsx"
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )
    except ImportError:
        # Fallback: CSV in WhatsApp-friendly format
        output = io.StringIO()
        import csv
        writer = csv.writer(output)
        writer.writerow(headers)
        for sheet_name, contacts in sheets.items():
            for c in contacts:
                phone_digits = "".join(ch for ch in str(c["phone"]) if ch.isdigit())
                if len(phone_digits) > 10:
                    phone_digits = phone_digits[-10:]
                writer.writerow([
                    c["name"], phone_digits, "91", c["whatsapp_number"],
                    sheet_name if category else ", ".join(c.get("categories", [])) or "General",
                    c.get("address", ""),
                ])
        output.seek(0)
        fname = f"whatsapp_contacts_{category or 'all'}_{date.today().isoformat()}.csv"
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )


# ── Rentals ──────────────────────────────────────────────────────────────

@app.route("/rentals")
@login_required
def rentals_list():
    status = request.args.get("status", "")
    query = Rental.query
    if status:
        query = query.filter_by(status=status)
    rentals = query.order_by(Rental.created_at.desc()).all()

    for rental in rentals:
        if rental.status == "active" and rental.is_overdue:
            rental.status = "overdue"
    db.session.commit()

    return render_template("rentals/list.html", rentals=rentals, current_status=status)


@app.route("/rentals/add", methods=["GET", "POST"])
@login_required
def rentals_add():
    if request.method == "POST":
        customer_id = int(request.form["customer_id"])
        start_date = date.fromisoformat(request.form["start_date"])
        end_date = date.fromisoformat(request.form["end_date"])
        item_ids = request.form.getlist("item_ids")

        if end_date < start_date:
            flash("End date must be after start date.", "error")
            return redirect(url_for("rentals_add"))
        if not item_ids:
            flash("Select at least one item.", "error")
            return redirect(url_for("rentals_add"))

        rental = Rental(
            rental_number=generate_number("RNT", Rental, "rental_number"),
            customer_id=customer_id,
            start_date=start_date,
            end_date=end_date,
            discount=float(request.form.get("discount", 0)),
            notes=request.form.get("notes", ""),
        )
        db.session.add(rental)
        db.session.flush()

        for item_id in item_ids:
            item = ClothingItem.query.get(int(item_id))
            if not item or item.status != "available":
                db.session.rollback()
                flash(f"Item '{item.name if item else item_id}' is not available.", "error")
                return redirect(url_for("rentals_add"))

            ri = RentalItem(
                rental_id=rental.id,
                item_id=item.id,
                daily_rate=item.daily_rate,
                deposit=item.deposit,
            )
            db.session.add(ri)
            item.status = "rented"

        rental.recalculate_totals()
        db.session.commit()

        invoice = Invoice(
            invoice_number=generate_number("INV", Invoice, "invoice_number"),
            rental_id=rental.id,
            due_date=end_date,
            subtotal=rental.total_amount,
            total=rental.total_amount,
        )
        db.session.add(invoice)
        db.session.commit()

        flash(f"Rental {rental.rental_number} created with invoice {invoice.invoice_number}.", "success")
        return redirect(url_for("rentals_view", id=rental.id))

    customers = Customer.query.order_by(Customer.name).all()
    available_items = ClothingItem.query.filter_by(status="available").order_by(ClothingItem.name).all()
    return render_template(
        "rentals/form.html",
        rental=None,
        customers=customers,
        available_items=available_items,
        today=date.today().isoformat(),
    )


@app.route("/rentals/<int:id>")
@login_required
def rentals_view(id):
    rental = Rental.query.get_or_404(id)
    invoice = rental.invoices.first()
    return render_template("rentals/view.html", rental=rental, invoice=invoice, today=date.today())


@app.route("/rentals/<int:id>/return", methods=["POST"])
@login_required
def rentals_return(id):
    rental = Rental.query.get_or_404(id)
    if rental.status in ("returned", "cancelled"):
        flash("Rental already closed.", "error")
        return redirect(url_for("rentals_view", id=id))

    return_date = date.fromisoformat(request.form.get("return_date", date.today().isoformat()))
    rental.actual_return_date = return_date
    rental.damage_fee = float(request.form.get("damage_fee", 0))
    rental.late_fee = float(request.form.get("late_fee", 0))

    if return_date > rental.end_date:
        extra_days = (return_date - rental.end_date).days
        daily_total = sum(ri.daily_rate for ri in rental.items)
        if rental.late_fee == 0:
            rental.late_fee = extra_days * daily_total * 0.5

    rental.recalculate_totals()
    rental.status = "returned"

    for ri in rental.items:
        ri.item.status = "available"

    invoice = rental.invoices.first()
    if invoice:
        invoice.subtotal = rental.total_amount
        invoice.total = rental.total_amount + invoice.tax_amount
        invoice.update_status()

    db.session.commit()
    flash(f"Rental {rental.rental_number} marked as returned.", "success")
    return redirect(url_for("rentals_view", id=id))


@app.route("/rentals/<int:id>/cancel", methods=["POST"])
@login_required
def rentals_cancel(id):
    rental = Rental.query.get_or_404(id)
    if rental.status != "active":
        flash("Only active rentals can be cancelled.", "error")
        return redirect(url_for("rentals_view", id=id))

    for ri in rental.items:
        ri.item.status = "available"
    rental.status = "cancelled"
    db.session.commit()
    flash("Rental cancelled.", "success")
    return redirect(url_for("rentals_list"))


# ── Invoices & Billing ──────────────────────────────────────────────────────

@app.route("/billing")
@login_required
def billing_list():
    status = request.args.get("status", "")
    query = Invoice.query
    if status:
        query = query.filter_by(status=status)
    invoices = query.order_by(Invoice.created_at.desc()).all()
    total_outstanding = sum(inv.balance_due for inv in invoices if inv.status != "paid")
    return render_template(
        "billing/list.html",
        invoices=invoices,
        current_status=status,
        total_outstanding=total_outstanding,
    )


@app.route("/billing/<int:id>")
def billing_view(id):
    invoice = Invoice.query.get_or_404(id)
    return render_template("billing/view.html", invoice=invoice, payment_methods=PAYMENT_METHODS)


@app.route("/billing/<int:id>/pay", methods=["POST"])
def billing_pay(id):
    invoice = Invoice.query.get_or_404(id)
    amount = float(request.form["amount"])
    if amount <= 0:
        flash("Payment amount must be positive.", "error")
        return redirect(url_for("billing_view", id=id))

    payment = Payment(
        invoice_id=invoice.id,
        amount=amount,
        method=request.form.get("method", "cash"),
        reference=request.form.get("reference", ""),
        notes=request.form.get("notes", ""),
    )
    db.session.add(payment)
    invoice.amount_paid += amount
    invoice.update_status()
    db.session.commit()
    flash(f"Payment of ₹{amount:,.2f} recorded.", "success")
    return redirect(url_for("billing_view", id=id))


@app.route("/billing/<int:id>/print")
def billing_print(id):
    invoice = Invoice.query.get_or_404(id)
    return render_template("billing/print.html", invoice=invoice)


# ── Reports ──────────────────────────────────────────────────────────────

@app.route("/reports")
def reports():
    today = date.today()
    month_start = today.replace(day=1)

    monthly_payments = Payment.query.filter(
        Payment.paid_at >= datetime.combine(month_start, datetime.min.time())
    ).all()
    monthly_revenue = sum(p.amount for p in monthly_payments)

    total_rentals = Rental.query.count()
    completed_rentals = Rental.query.filter_by(status="returned").count()

    popular = (
        db.session.query(ClothingItem.name, db.func.count(RentalItem.id).label("times"))
        .join(RentalItem, RentalItem.item_id == ClothingItem.id)
        .group_by(ClothingItem.id)
        .order_by(db.desc("times"))
        .limit(5)
        .all()
    )

    revenue_by_method = (
        db.session.query(Payment.method, db.func.sum(Payment.amount))
        .filter(Payment.paid_at >= datetime.combine(month_start, datetime.min.time()))
        .group_by(Payment.method)
        .all()
    )

    return render_template(
        "reports.html",
        monthly_revenue=monthly_revenue,
        total_rentals=total_rentals,
        completed_rentals=completed_rentals,
        popular=popular,
        revenue_by_method=revenue_by_method,
        month_name=month_start.strftime("%B %Y"),
    )


# ── Booking Panel ──────────────────────────────────────────────────────────────

@app.route("/booking")
@login_required
def booking_panel():
    bookings = Booking.query.filter(Booking.status != "cancelled").order_by(Booking.created_at.desc()).all()
    return render_template("booking/panel.html", bookings=bookings, today=date.today())


@app.route("/booking/new", methods=["GET", "POST"])
@login_required
def booking_new():
    if request.method == "POST":
        customer_name = request.form["customer_name"].strip()
        customer_address = request.form["customer_address"].strip()
        contact_1 = request.form["contact_1"].strip()
        whatsapp_no = request.form.get("whatsapp_no", "").strip()
        delivery_date = date.fromisoformat(request.form["delivery_date"])
        delivery_time = request.form["delivery_time"].strip()
        return_date = date.fromisoformat(request.form["return_date"])
        return_time = request.form["return_time"].strip()
        venue = request.form.get("venue", "").strip()
        security_deposit = float(request.form.get("security_deposit", 0))
        common_notes = request.form.get("common_notes", "").strip()
        staff_names_list = request.form.getlist("staff_names[]")
        staff_names = ", ".join(staff_names_list) if staff_names_list else ""

        if return_date < delivery_date:
            flash("Return date must be after delivery date.", "error")
            return redirect(url_for("booking_new"))

        # Parse multiple dresses
        item_ids = request.form.getlist("item_ids[]")
        dress_names = request.form.getlist("dress_names[]")
        prices = request.form.getlist("prices[]")
        advances = request.form.getlist("advances[]")
        dress_notes = request.form.getlist("dress_notes[]")

        if not item_ids:
            flash("Please select at least one dress.", "error")
            return redirect(url_for("booking_new"))

        # Validate all dresses and check for conflicts
        items_to_book = []
        for i, raw_id in enumerate(item_ids):
            item_id = int(raw_id)
            item = ClothingItem.query.get(item_id)
            if not item:
                flash(f"Dress '{dress_names[i]}' not found.", "error")
                return redirect(url_for("booking_new"))

            # Prevent double-booking; edge-day cases (returning on our delivery / booked on our return) are allowed
            conflict = Booking.query.filter(
                Booking.status.in_(["booked", "delivered"]),
                Booking.delivery_date < return_date,
                Booking.return_date > delivery_date,
                Booking.return_date != delivery_date,
                Booking.delivery_date != return_date,
            ).join(BookingItem).filter(BookingItem.item_id == item_id).first()

            if not conflict:
                conflict = Booking.query.filter(
                    Booking.item_id == item_id,
                    Booking.status.in_(["booked", "delivered"]),
                    Booking.delivery_date < return_date,
                    Booking.return_date > delivery_date,
                    Booking.return_date != delivery_date,
                    Booking.delivery_date != return_date,
                ).first()

            if conflict:
                flash(
                    f"'{dress_names[i]}' is already booked from "
                    f"{conflict.delivery_date.strftime('%d %b %Y')} to "
                    f"{conflict.return_date.strftime('%d %b %Y')} "
                    f"(Serial #{conflict.monthly_serial:02d}). ",
                    "error"
                )
                return redirect(url_for("booking_new"))

            items_to_book.append(item)

        booking_number = generate_number("BKG", Booking, "booking_number")

        # Monthly serial based on DELIVERY month (not booking date)
        # Skips numbers whose digit sum is 4 or 8
        del_month_start = delivery_date.replace(day=1)
        if delivery_date.month == 12:
            del_month_end = date(delivery_date.year + 1, 1, 1)
        else:
            del_month_end = date(delivery_date.year, delivery_date.month + 1, 1)
        month_booking_count = Booking.query.filter(
            Booking.delivery_date >= del_month_start,
            Booking.delivery_date < del_month_end,
        ).count()
        # Convert position (count+1) to the actual valid serial number
        monthly_serial = serial_position_to_value(month_booking_count + 1)

        # Calculate totals
        total_price = sum(float(p) for p in prices)
        total_advance = sum(float(a) for a in advances)
        total_remaining = total_price - total_advance

        booking = Booking(
            booking_number=booking_number,
            monthly_serial=monthly_serial,
            customer_name=customer_name,
            customer_address=customer_address,
            contact_1=contact_1,
            whatsapp_no=whatsapp_no,
            venue=venue,
            staff_names=staff_names,
            delivery_date=delivery_date,
            delivery_time=delivery_time,
            return_date=return_date,
            return_time=return_time,
            security_deposit=security_deposit,
            total_price=total_price,
            total_advance=total_advance,
            total_remaining=total_remaining,
            common_notes=common_notes,
            # Legacy fields for first dress
            item_id=int(item_ids[0]),
            dress_name=dress_names[0],
            price=total_price,
            advance=total_advance,
            remaining=total_remaining,
        )
        db.session.add(booking)
        db.session.flush()

        # Add each dress as BookingItem
        for i, item in enumerate(items_to_book):
            d_price = float(prices[i])
            d_advance = float(advances[i])
            d_note = dress_notes[i] if i < len(dress_notes) else ""
            bi = BookingItem(
                booking_id=booking.id,
                item_id=item.id,
                dress_name=dress_names[i],
                category=item.category,
                size=item.size or "",
                price=d_price,
                advance=d_advance,
                remaining=d_price - d_advance,
                notes=d_note,
            )
            db.session.add(bi)
            item.status = "rented"

        # Auto-add customer to Customers table
        existing_customer = Customer.query.filter_by(phone=contact_1).first()
        if not existing_customer:
            new_cust = Customer(
                name=customer_name,
                phone=contact_1,
                email="",
                address=customer_address,
            )
            db.session.add(new_cust)

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Error creating booking: {e}", "error")
            return redirect(url_for("booking_new"))

        flash(f"Booking Serial #{monthly_serial:02d} created ({len(items_to_book)} dresses)! Remaining: ₹{total_remaining:,.0f}", "success")
        if request.form.get("print_after_save") == "1":
            return redirect(url_for("booking_print", id=booking.id))
        return redirect(url_for("booking_view", id=booking.id))

    staff_list = Staff.query.filter_by(active=True).order_by(Staff.name).all()
    return render_template("booking/form.html", today=date.today().isoformat(), categories=CATEGORIES, staff_list=staff_list)


@app.route("/booking/<int:id>")
@login_required
def booking_view(id):
    booking = Booking.query.get_or_404(id)
    return render_template("booking/view.html", booking=booking, today=date.today())


@app.route("/booking/<int:id>/edit", methods=["GET", "POST"])
@login_required
def booking_edit(id):
    booking = Booking.query.get_or_404(id)

    if request.method == "POST":
        customer_name = request.form["customer_name"].strip()
        customer_address = request.form["customer_address"].strip()
        contact_1 = request.form["contact_1"].strip()
        whatsapp_no = request.form.get("whatsapp_no", "").strip()
        delivery_date = date.fromisoformat(request.form["delivery_date"])
        delivery_time = request.form["delivery_time"].strip()
        return_date = date.fromisoformat(request.form["return_date"])
        return_time = request.form["return_time"].strip()
        venue = request.form.get("venue", "").strip()
        security_deposit = float(request.form.get("security_deposit", 0) or 0)
        common_notes = request.form.get("common_notes", "").strip()
        staff_names_list = request.form.getlist("staff_names[]")
        staff_names = ", ".join(staff_names_list) if staff_names_list else ""

        if return_date < delivery_date:
            flash("Return date must be on or after delivery date.", "error")
            return redirect(url_for("booking_edit", id=id))

        item_ids = request.form.getlist("item_ids[]")
        dress_names = request.form.getlist("dress_names[]")
        prices = request.form.getlist("prices[]")
        advances = request.form.getlist("advances[]")
        dress_notes = request.form.getlist("dress_notes[]")

        if not item_ids:
            flash("Please select at least one dress.", "error")
            return redirect(url_for("booking_edit", id=id))

        # Collect old item ids before replacing so we can free them
        old_item_ids = set()
        for bi in booking.booking_items:
            old_item_ids.add(bi.item_id)
        if booking.item_id:
            old_item_ids.add(booking.item_id)

        # Validate all dresses and check for strict conflicts (excluding THIS booking)
        # Edge-day bookings (returning on our delivery date / booked on our return date) are ALLOWED
        items_to_book = []
        new_item_ids = set()
        for i, raw_id in enumerate(item_ids):
            item_id = int(raw_id)
            item = ClothingItem.query.get(item_id)
            if not item:
                flash(f"Dress '{dress_names[i]}' not found.", "error")
                return redirect(url_for("booking_edit", id=id))

            conflict = Booking.query.filter(
                Booking.id != booking.id,
                Booking.status.in_(["booked", "delivered"]),
                Booking.delivery_date < return_date,
                Booking.return_date > delivery_date,
                Booking.return_date != delivery_date,    # allow: other returns on our delivery day
                Booking.delivery_date != return_date,   # allow: other delivers on our return day
            ).join(BookingItem).filter(BookingItem.item_id == item_id).first()

            if not conflict:
                conflict = Booking.query.filter(
                    Booking.id != booking.id,
                    Booking.item_id == item_id,
                    Booking.status.in_(["booked", "delivered"]),
                    Booking.delivery_date < return_date,
                    Booking.return_date > delivery_date,
                    Booking.return_date != delivery_date,
                    Booking.delivery_date != return_date,
                ).first()

            if conflict:
                flash(
                    f"'{dress_names[i]}' is already booked from "
                    f"{conflict.delivery_date.strftime('%d %b %Y')} to "
                    f"{conflict.return_date.strftime('%d %b %Y')} "
                    f"(Serial #{conflict.monthly_serial:02d}). Choose different dates or dress.",
                    "error"
                )
                return redirect(url_for("booking_edit", id=id))

            items_to_book.append(item)
            new_item_ids.add(item_id)

        # Totals
        total_price = sum(float(p or 0) for p in prices)
        total_advance = sum(float(a or 0) for a in advances)
        total_remaining = total_price - total_advance

        # Update booking core fields
        booking.customer_name = customer_name
        booking.customer_address = customer_address
        booking.contact_1 = contact_1
        booking.whatsapp_no = whatsapp_no
        booking.venue = venue
        booking.staff_names = staff_names
        booking.delivery_date = delivery_date
        booking.delivery_time = delivery_time
        booking.return_date = return_date
        booking.return_time = return_time
        booking.security_deposit = security_deposit
        booking.common_notes = common_notes
        booking.total_price = total_price
        booking.total_advance = total_advance
        booking.total_remaining = total_remaining
        # Legacy mirror fields (first dress)
        booking.item_id = int(item_ids[0])
        booking.dress_name = dress_names[0]
        booking.price = total_price
        booking.advance = total_advance
        booking.remaining = total_remaining

        # Replace booking items
        for bi in list(booking.booking_items):
            db.session.delete(bi)
        db.session.flush()

        # Free any items that are no longer in this booking
        freed_ids = old_item_ids - new_item_ids
        for freed_id in freed_ids:
            freed_item = ClothingItem.query.get(freed_id)
            if freed_item:
                # Only free if not booked in any other active booking
                still_booked = Booking.query.filter(
                    Booking.id != booking.id,
                    Booking.status.in_(["booked", "delivered"]),
                ).join(BookingItem).filter(BookingItem.item_id == freed_id).first()
                if not still_booked:
                    freed_item.status = "available"

        for i, item in enumerate(items_to_book):
            d_price = float(prices[i] or 0)
            d_advance = float(advances[i] or 0)
            d_note = dress_notes[i] if i < len(dress_notes) else ""
            db.session.add(BookingItem(
                booking_id=booking.id,
                item_id=item.id,
                dress_name=dress_names[i],
                category=item.category,
                size=item.size or "",
                price=d_price,
                advance=d_advance,
                remaining=d_price - d_advance,
                notes=d_note,
            ))
            item.status = "rented"

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Error updating booking: {e}", "error")
            return redirect(url_for("booking_edit", id=id))

        flash(f"Booking Serial #{booking.monthly_serial:02d} updated successfully!", "success")
        if request.form.get("print_after_save") == "1":
            return redirect(url_for("booking_print", id=booking.id))
        return redirect(url_for("booking_view", id=booking.id))

    staff_list = Staff.query.filter_by(active=True).order_by(Staff.name).all()
    return render_template("booking/form.html", edit_booking=booking, staff_list=staff_list)


@app.route("/booking/<int:id>/print")
@login_required
def booking_print(id):
    booking = Booking.query.get_or_404(id)
    total = booking.total_price or booking.price
    # GST 18% inclusive: base = total / 1.18, sgst = cgst = base * 0.09
    base_amount = round(total / 1.18, 2)
    sgst = round(base_amount * 0.09, 2)
    cgst = round(base_amount * 0.09, 2)
    return render_template("booking/print_bill.html", booking=booking, base_amount=base_amount, sgst=sgst, cgst=cgst)


@app.route("/booking/<int:id>/deliver", methods=["GET", "POST"])
@login_required
def booking_deliver(id):
    """Redirect to delivery panel - bookings can only be marked delivered from there."""
    return redirect(url_for("booking_delivery_detail", booking_id=id))


@app.route("/booking/<int:id>/return", methods=["GET", "POST"])
@login_required
def booking_return(id):
    """Redirect to return panel - bookings can only be marked returned from there."""
    return redirect(url_for("booking_return_detail", booking_id=id))


@app.route("/booking/<int:id>/cancel", methods=["POST"])
@owner_required
def booking_cancel(id):
    booking = Booking.query.get_or_404(id)
    if booking.status in ("returned", "cancelled"):
        flash("Cannot cancel this booking.", "error")
        return redirect(url_for("booking_view", id=id))
    booking.status = "cancelled"
    # Free all items in this booking
    for bi in booking.booking_items:
        item = ClothingItem.query.get(bi.item_id)
        if item:
            item.status = "available"
    if booking.item_id and booking.item:
        booking.item.status = "available"
    db.session.commit()
    flash(f"Booking Serial #{booking.monthly_serial:02d} cancelled.", "success")
    return redirect(url_for("booking_panel"))


@app.route("/api/booking/next-serial")
def api_booking_next_serial():
    """Get the next serial number for a given delivery month."""
    dd_str = request.args.get("delivery_date", "")
    try:
        dd = date.fromisoformat(dd_str)
    except ValueError:
        dd = date.today()
    month_start = dd.replace(day=1)
    if dd.month == 12:
        month_end = date(dd.year + 1, 1, 1)
    else:
        month_end = date(dd.year, dd.month + 1, 1)
    count = Booking.query.filter(
        Booking.delivery_date >= month_start,
        Booking.delivery_date < month_end,
    ).count()
    return jsonify({"serial": serial_position_to_value(count + 1)})


@app.route("/api/booking/date-check")
@login_required
def api_booking_date_check():
    """
    Live conflict check for the edit form.
    Given a booking_id + new delivery/return dates + item_ids,
    returns per-item status: 'ok' | 'hard_conflict' | 'returning_warning' | 'booked_on_return_warning'.
    Hard conflict = another booking strictly overlaps → BLOCK.
    Edge warnings = dress returning on delivery date OR booked on return date → ALLOW with warning.
    """
    try:
        booking_id     = int(request.args.get("booking_id", 0))
        d_str          = request.args.get("delivery_date", "")
        r_str          = request.args.get("return_date", "")
        item_ids_raw   = request.args.getlist("item_ids[]")
    except (ValueError, TypeError):
        return jsonify({"error": "bad params"}), 400

    if not d_str or not r_str or not item_ids_raw:
        return jsonify([])

    try:
        d_date = date.fromisoformat(d_str)
        r_date = date.fromisoformat(r_str)
    except ValueError:
        return jsonify({"error": "bad dates"}), 400

    if r_date < d_date:
        return jsonify({"error": "return before delivery"}), 400

    results = []
    for raw_id in item_ids_raw:
        try:
            item_id = int(raw_id)
        except ValueError:
            continue

        item = ClothingItem.query.get(item_id)
        if not item:
            continue

        # ── HARD CONFLICT: strict overlap (not including same-day edge cases) ──
        hard = (
            Booking.query
            .filter(
                Booking.id != booking_id,
                Booking.status.in_(["booked", "delivered"]),
                Booking.delivery_date < r_date,   # other starts before this ends
                Booking.return_date > d_date,      # other ends after this starts
                Booking.return_date != d_date,     # exclude: returns exactly on our delivery (edge)
                Booking.delivery_date != r_date,   # exclude: delivers exactly on our return (edge)
            )
            .join(BookingItem, BookingItem.booking_id == Booking.id)
            .filter(BookingItem.item_id == item_id)
            .first()
        )
        # Also check legacy item_id field
        if not hard:
            hard = (
                Booking.query
                .filter(
                    Booking.id != booking_id,
                    Booking.item_id == item_id,
                    Booking.status.in_(["booked", "delivered"]),
                    Booking.delivery_date < r_date,
                    Booking.return_date > d_date,
                    Booking.return_date != d_date,
                    Booking.delivery_date != r_date,
                )
                .first()
            )

        if hard:
            results.append({
                "item_id": item_id,
                "item_name": item.name,
                "status": "hard_conflict",
                "conflict": {
                    "customer": hard.customer_name,
                    "serial":   hard.monthly_serial,
                    "delivery_date": hard.delivery_date.strftime("%d %b %Y"),
                    "return_date":   hard.return_date.strftime("%d %b %Y"),
                    "delivery_time": hard.delivery_time,
                    "return_time":   hard.return_time,
                    "venue":   hard.venue or "",
                    "contact": hard.contact_1 or "",
                },
            })
            continue

        # ── EDGE WARNING 1: another booking returns exactly on our delivery date ──
        ret_warn_b = (
            Booking.query
            .filter(
                Booking.id != booking_id,
                Booking.status.in_(["booked", "delivered"]),
                Booking.return_date == d_date,
            )
            .join(BookingItem, BookingItem.booking_id == Booking.id)
            .filter(BookingItem.item_id == item_id)
            .first()
        )
        if not ret_warn_b:
            ret_warn_b = Booking.query.filter(
                Booking.id != booking_id,
                Booking.item_id == item_id,
                Booking.status.in_(["booked", "delivered"]),
                Booking.return_date == d_date,
            ).first()

        if ret_warn_b:
            results.append({
                "item_id": item_id,
                "item_name": item.name,
                "status": "returning_warning",
                "conflict": {
                    "customer": ret_warn_b.customer_name,
                    "serial":   ret_warn_b.monthly_serial,
                    "delivery_date": ret_warn_b.delivery_date.strftime("%d %b %Y"),
                    "return_date":   ret_warn_b.return_date.strftime("%d %b %Y"),
                    "return_time":   ret_warn_b.return_time,
                    "venue":   ret_warn_b.venue or "",
                    "contact": ret_warn_b.contact_1 or "",
                },
            })
            continue

        # ── EDGE WARNING 2: another booking delivers exactly on our return date ──
        del_warn_b = (
            Booking.query
            .filter(
                Booking.id != booking_id,
                Booking.status.in_(["booked", "delivered"]),
                Booking.delivery_date == r_date,
            )
            .join(BookingItem, BookingItem.booking_id == Booking.id)
            .filter(BookingItem.item_id == item_id)
            .first()
        )
        if not del_warn_b:
            del_warn_b = Booking.query.filter(
                Booking.id != booking_id,
                Booking.item_id == item_id,
                Booking.status.in_(["booked", "delivered"]),
                Booking.delivery_date == r_date,
            ).first()

        if del_warn_b:
            results.append({
                "item_id": item_id,
                "item_name": item.name,
                "status": "booked_on_return_warning",
                "conflict": {
                    "customer": del_warn_b.customer_name,
                    "serial":   del_warn_b.monthly_serial,
                    "delivery_date": del_warn_b.delivery_date.strftime("%d %b %Y"),
                    "return_date":   del_warn_b.return_date.strftime("%d %b %Y"),
                    "delivery_time": del_warn_b.delivery_time,
                    "venue":   del_warn_b.venue or "",
                    "contact": del_warn_b.contact_1 or "",
                },
            })
            continue

        # ── All clear ──
        results.append({"item_id": item_id, "item_name": item.name, "status": "ok"})

    return jsonify(results)


@app.route("/api/booking/available-items")
def api_booking_available_items():
    """Get available items including those returning on delivery date and booked on return date (with warnings)."""
    delivery_date_str = request.args.get("delivery_date", "")
    return_date_str = request.args.get("return_date", "")
    try:
        exclude_booking = int(request.args.get("exclude_booking", 0)) or None
    except ValueError:
        exclude_booking = None

    if not delivery_date_str or not return_date_str:
        return jsonify({"free_items": [], "returning_items": [], "booked_on_return": []})

    try:
        d_date = date.fromisoformat(delivery_date_str)
        r_date = date.fromisoformat(return_date_str)
    except ValueError:
        return jsonify({"free_items": [], "returning_items": [], "booked_on_return": []})

    all_items = ClothingItem.query.filter(
        ClothingItem.status != "maintenance"
    ).all()

    # Overlapping bookings (strictly between dates, excluding edge cases)
    overlapping_q = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.delivery_date <= r_date,
        Booking.return_date >= d_date,
    )
    if exclude_booking:
        overlapping_q = overlapping_q.filter(Booking.id != exclude_booking)
    overlapping_bookings = overlapping_q.all()
    booked_item_ids = set()
    for b in overlapping_bookings:
        if b.booking_items:
            for bi in b.booking_items:
                booked_item_ids.add(bi.item_id)
        elif b.item_id:
            booked_item_ids.add(b.item_id)

    # Overlapping rentals
    overlapping_rentals = Rental.query.filter(
        Rental.status.in_(["active", "overdue"]),
        Rental.start_date <= r_date,
        Rental.end_date >= d_date,
    ).all()
    rented_item_ids = set()
    for rental in overlapping_rentals:
        for ri in rental.items:
            rented_item_ids.add(ri.item_id)

    # Items returning on delivery date (these ARE bookable with warning)
    returning_q = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.return_date == d_date,
    )
    if exclude_booking:
        returning_q = returning_q.filter(Booking.id != exclude_booking)
    returning_on_delivery_bookings = returning_q.all()
    returning_on_delivery_ids = set()
    returning_info = {}
    for b in returning_on_delivery_bookings:
        items_in_b = []
        if b.booking_items:
            items_in_b = [bi.item_id for bi in b.booking_items]
        elif b.item_id:
            items_in_b = [b.item_id]
        for bid in items_in_b:
            returning_on_delivery_ids.add(bid)
            returning_info[bid] = {
                "return_time": b.return_time,
                "customer": b.customer_name,
                "booking_number": b.booking_number,
                "serial_no": b.monthly_serial,
                "total_rent": b.total_price or b.price,
                "venue": b.venue or "",
            }

    # Items booked for delivery on return date (these ARE bookable with warning)
    booked_on_return_q = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.delivery_date == r_date,
    )
    if exclude_booking:
        booked_on_return_q = booked_on_return_q.filter(Booking.id != exclude_booking)
    bookings_on_return = booked_on_return_q.all()
    booked_on_return_ids = set()
    booked_on_return_info = {}
    for b in bookings_on_return:
        items_in_b = []
        if b.booking_items:
            items_in_b = [bi.item_id for bi in b.booking_items]
        elif b.item_id:
            items_in_b = [b.item_id]
        for bid in items_in_b:
            booked_on_return_ids.add(bid)
            booked_on_return_info[bid] = {
                "delivery_time": b.delivery_time,
                "customer": b.customer_name,
                "booking_number": b.booking_number,
                "serial_no": b.monthly_serial,
                "total_rent": b.total_price or b.price,
                "venue": b.venue or "",
            }

    # Truly busy: overlapping but NOT edge-case items (returning on delivery / booked on return)
    busy_ids = (booked_item_ids | rented_item_ids) - returning_on_delivery_ids - booked_on_return_ids

    free_items = []
    for i in all_items:
        if i.id not in busy_ids:
            item_data = {
                "id": i.id, "name": i.name,
                "display_name": dress_display_name(i.name, i.category, i.size),
                "sku": i.sku, "category": i.category,
                "color": i.color, "size": i.size, "item_type": i.item_type,
                "sub_category": i.sub_category or "Normal",
                "photo": i.photo or "",
                "returning_warning": returning_info.get(i.id),
                "booked_warning": booked_on_return_info.get(i.id),
            }
            free_items.append(item_data)

    # Full details for returning items section
    returning_items = []
    seen_ids = set()
    for b in returning_on_delivery_bookings:
        items_in_b = []
        if b.booking_items:
            items_in_b = [(bi.item_id, bi.dress_name) for bi in b.booking_items]
        elif b.item_id:
            items_in_b = [(b.item_id, b.dress_name)]
        for bid, bname in items_in_b:
            if bid not in seen_ids:
                returning_items.append({
                    "id": bid, "name": bname,
                    "return_time": b.return_time, "customer": b.customer_name,
                    "booking_number": b.booking_number, "serial_no": b.monthly_serial,
                    "total_rent": b.total_price or b.price, "venue": b.venue or "",
                })
                seen_ids.add(bid)

    # Full details for booked on return section
    booked_on_return = []
    for b in bookings_on_return:
        items_in_b = []
        if b.booking_items:
            items_in_b = [(bi.item_id, bi.dress_name) for bi in b.booking_items]
        elif b.item_id:
            items_in_b = [(b.item_id, b.dress_name)]
        for bid, bname in items_in_b:
            item_obj = ClothingItem.query.get(bid)
            booked_on_return.append({
                "id": bid, "name": bname,
                "category": item_obj.category if item_obj else "",
                "customer": b.customer_name, "booking_number": b.booking_number,
                "serial_no": b.monthly_serial, "delivery_time": b.delivery_time,
                "total_rent": b.total_price or b.price, "venue": b.venue or "",
                "photo": item_obj.photo or "" if item_obj else "",
            })

    return jsonify({
        "free_items": free_items,
        "returning_items": returning_items,
        "booked_on_return": booked_on_return,
    })


def _booking_uses_item(booking, item_id):
    if booking.booking_items:
        return any(bi.item_id == item_id for bi in booking.booking_items)
    return booking.item_id == item_id


def _serialize_booking_conflict(b):
    return {
        "customer": b.customer_name,
        "serial_no": b.monthly_serial,
        "delivery_date": b.delivery_date.strftime("%d %b %Y"),
        "delivery_time": b.delivery_time,
        "return_date": b.return_date.strftime("%d %b %Y"),
        "return_time": b.return_time,
        "venue": b.venue or "",
        "total_rent": b.total_price or b.price,
        "contact": b.contact_1 or "",
        "booking_id": b.id,
    }


def check_item_availability_for_dates(item, d_date, r_date, exclude_booking=None):
    """Return availability status for one inventory item between delivery and return dates."""
    if item.status == "maintenance":
        return {
            "status": "not_available",
            "reason": "Item is under maintenance",
            "returning_warning": None,
            "booked_warning": None,
            "blocking_booking": None,
        }

    overlapping_q = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.delivery_date <= r_date,
        Booking.return_date >= d_date,
    )
    if exclude_booking:
        overlapping_q = overlapping_q.filter(Booking.id != exclude_booking)
    overlapping_bookings = overlapping_q.all()

    returning_warning = None
    booked_warning = None
    blocking_booking = None

    for b in overlapping_bookings:
        if not _booking_uses_item(b, item.id):
            continue
        if b.return_date == d_date:
            returning_warning = _serialize_booking_conflict(b)
            returning_warning["return_time"] = b.return_time
            continue
        if b.delivery_date == r_date:
            booked_warning = _serialize_booking_conflict(b)
            continue
        blocking_booking = _serialize_booking_conflict(b)
        break

    rental_block = (
        Rental.query.filter(
            Rental.status.in_(["active", "overdue"]),
            Rental.start_date <= r_date,
            Rental.end_date >= d_date,
        )
        .join(RentalItem, RentalItem.rental_id == Rental.id)
        .filter(RentalItem.item_id == item.id)
        .first()
    )
    if rental_block and not blocking_booking:
        blocking_booking = {
            "customer": rental_block.customer.name if rental_block.customer else "Rental",
            "serial_no": rental_block.rental_number,
            "delivery_date": rental_block.start_date.strftime("%d %b %Y"),
            "return_date": rental_block.end_date.strftime("%d %b %Y"),
            "venue": "",
            "total_rent": rental_block.total_amount,
            "contact": "",
            "booking_id": None,
            "type": "rental",
        }

    if blocking_booking:
        return {
            "status": "not_available",
            "reason": "Booked during selected dates",
            "returning_warning": returning_warning,
            "booked_warning": booked_warning,
            "blocking_booking": blocking_booking,
        }

    if returning_warning or booked_warning:
        return {
            "status": "available_with_warning",
            "reason": "Available with scheduling note",
            "returning_warning": returning_warning,
            "booked_warning": booked_warning,
            "blocking_booking": None,
        }

    return {
        "status": "available",
        "reason": "Free for entire period",
        "returning_warning": None,
        "booked_warning": None,
        "blocking_booking": None,
    }


@app.route("/api/dress-checker")
@login_required
def api_dress_checker():
    """Check if a dress (by name) is free between delivery and return dates."""
    delivery_date_str = request.args.get("delivery_date", "")
    return_date_str = request.args.get("return_date", "")
    dress_name = request.args.get("dress_name", "").strip()
    category_filter = request.args.get("category", "").strip()

    if not delivery_date_str or not return_date_str:
        return jsonify({"error": "Delivery and return dates are required"}), 400
    if not dress_name:
        return jsonify({"error": "Dress name is required"}), 400

    try:
        d_date = date.fromisoformat(delivery_date_str)
        r_date = date.fromisoformat(return_date_str)
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400

    if r_date < d_date:
        return jsonify({"error": "Return date cannot be before delivery date"}), 400

    query = ClothingItem.query
    name_filter = dress_name_sql_filter(dress_name)
    if name_filter is not None:
        query = query.filter(name_filter)
    if category_filter:
        query = query.filter(ClothingItem.category == category_filter)
    items = query.order_by(ClothingItem.name, ClothingItem.size).all()

    if not items:
        return jsonify({
            "items": [],
            "message": f"No dress found matching '{dress_name}'" + (f" in {category_filter}" if category_filter else ""),
            "delivery_date": d_date.strftime("%d %b %Y"),
            "return_date": r_date.strftime("%d %b %Y"),
        })

    results = []
    for item in items:
        avail = check_item_availability_for_dates(item, d_date, r_date)
        results.append({
            "id": item.id,
            "name": item.name,
            "display_name": dress_display_name(item.name, item.category, item.size),
            "sku": item.sku,
            "category": item.category,
            "size": item.size or "",
            "color": item.color or "",
            "photo": item.photo or "",
            "inventory_status": item.status,
            **avail,
        })

    return jsonify({
        "items": results,
        "message": f"Found {len(results)} matching dress(es)",
        "delivery_date": d_date.strftime("%d %b %Y"),
        "return_date": r_date.strftime("%d %b %Y"),
        "dress_name": dress_name,
        "category": category_filter or "All",
    })


# ── API helpers ──────────────────────────────────────────────────────────────

@app.route("/api/dashboard/free-items")
def api_dashboard_free_items():
    """Dashboard free-item finder with warnings for items booked on return date."""
    delivery_date_str = request.args.get("delivery_date", "")
    return_date_str = request.args.get("return_date", "")
    category_filter = request.args.get("category", "")

    if not delivery_date_str or not return_date_str:
        return jsonify({"free_items": [], "returning_on_delivery": [], "warnings": {}})

    try:
        d_date = date.fromisoformat(delivery_date_str)
        r_date = date.fromisoformat(return_date_str)
    except ValueError:
        return jsonify({"free_items": [], "returning_on_delivery": [], "warnings": {}})

    all_items_query = ClothingItem.query.filter(ClothingItem.status != "maintenance")
    if category_filter:
        all_items_query = all_items_query.filter(ClothingItem.category == category_filter)
    all_items = all_items_query.all()

    # Items booked/delivered overlapping with requested period
    overlapping_bookings = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.delivery_date <= r_date,
        Booking.return_date >= d_date,
    ).all()
    booked_item_ids = set()
    for b in overlapping_bookings:
        if b.booking_items:
            for bi in b.booking_items:
                booked_item_ids.add(bi.item_id)
        elif b.item_id:
            booked_item_ids.add(b.item_id)

    # Items in active rentals overlapping
    overlapping_rentals = Rental.query.filter(
        Rental.status.in_(["active", "overdue"]),
        Rental.start_date <= r_date,
        Rental.end_date >= d_date,
    ).all()
    rented_item_ids = set()
    for rental in overlapping_rentals:
        for ri in rental.items:
            rented_item_ids.add(ri.item_id)

    busy_ids = booked_item_ids | rented_item_ids

    # Items returning ON delivery date - exclude from free list
    returning_on_delivery_bookings = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.return_date == d_date,
    ).all()
    returning_on_delivery_ids = set()
    for b in returning_on_delivery_bookings:
        if b.booking_items:
            for bi in b.booking_items:
                returning_on_delivery_ids.add(bi.item_id)
        elif b.item_id:
            returning_on_delivery_ids.add(b.item_id)

    # Items booked for delivery on return date
    bookings_on_return_date = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.delivery_date == r_date,
    ).all()
    booked_on_return_ids = set()
    booked_on_return_info = {}
    for b in bookings_on_return_date:
        items_in_b = []
        if b.booking_items:
            items_in_b = [bi.item_id for bi in b.booking_items]
        elif b.item_id:
            items_in_b = [b.item_id]
        for bid in items_in_b:
            booked_on_return_ids.add(bid)
            booked_on_return_info[bid] = {
                "booking_number": b.booking_number,
                "serial_no": b.monthly_serial,
                "customer_name": b.customer_name,
                "amount": b.total_price or b.price,
                "delivery_time": b.delivery_time,
                "venue": b.venue or "",
            }

    # Truly busy: overlapping but NOT edge-case items
    truly_busy = busy_ids - returning_on_delivery_ids - booked_on_return_ids

    # Free items (includes edge-case items with warnings)
    free_items = []
    warnings = {}
    for i in all_items:
        if i.id not in truly_busy:
            free_items.append({
                "id": i.id, "name": i.name,
                "display_name": dress_display_name(i.name, i.category, i.size),
                "category": i.category,
                "color": i.color or "", "size": i.size or "",
            })
            if i.id in booked_on_return_ids:
                warnings[str(i.id)] = booked_on_return_info[i.id]

    # Returning on delivery date - full booking details
    returning_on_delivery = []
    for b in returning_on_delivery_bookings:
        items_in_b = []
        if b.booking_items:
            items_in_b = [(bi.item_id, bi.dress_name, bi.category) for bi in b.booking_items]
        elif b.item_id:
            items_in_b = [(b.item_id, b.dress_name, "")]

        for bid, bname, bcat in items_in_b:
            item = ClothingItem.query.get(bid)
            cat = bcat or (item.category if item else "")
            if category_filter and cat != category_filter:
                continue
            returning_on_delivery.append({
                "id": bid,
                "dress_name": bname,
                "display_name": dress_display_name(bname, cat, item.size if item else ""),
                "category": cat,
                "booking_number": b.booking_number,
                "serial_no": b.monthly_serial,
                "customer_name": b.customer_name,
                "contact": b.contact_1,
                "amount": b.total_price or b.price,
                "return_time": b.return_time,
                "return_date": b.return_date.strftime("%d %b %Y"),
                "venue": b.venue or "",
            })

    return jsonify({
        "free_items": free_items,
        "returning_on_delivery": returning_on_delivery,
        "warnings": warnings,
    })


@app.route("/api/items/available")
def api_available_items():
    items = ClothingItem.query.filter_by(status="available").all()
    return jsonify([
        {"id": i.id, "name": i.name, "sku": i.sku, "daily_rate": i.daily_rate, "deposit": i.deposit}
        for i in items
    ])


# ── CSV Export ───────────────────────────────────────────────────────────────

@app.route("/admin/export/bookings.csv")
@owner_required
def export_bookings_csv():
    """Download all bookings as a CSV file."""
    import csv, io
    from flask import Response
    bookings = Booking.query.filter(
        Booking.status != "cancelled"
    ).order_by(Booking.delivery_date.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Serial#", "Booking#", "Status", "Customer", "Address", "Contact", "WhatsApp",
        "Venue", "Delivery Date", "Delivery Time", "Return Date", "Return Time",
        "Dresses", "Total Rent", "Advance Paid", "Remaining", "Security Deposit",
        "Common Notes", "Staff", "Created At"
    ])
    for b in bookings:
        if b.booking_items:
            dresses = " | ".join(bi.dress_name for bi in b.booking_items)
        else:
            dresses = b.dress_name or ""
        writer.writerow([
            f"#{b.monthly_serial:02d}" if b.monthly_serial else "",
            b.booking_number or "",
            b.status,
            b.customer_name,
            b.customer_address or "",
            b.contact_1 or "",
            b.whatsapp_no or "",
            b.venue or "",
            b.delivery_date.strftime("%d-%m-%Y") if b.delivery_date else "",
            b.delivery_time or "",
            b.return_date.strftime("%d-%m-%Y") if b.return_date else "",
            b.return_time or "",
            dresses,
            b.total_price or b.price or 0,
            b.total_advance or b.advance or 0,
            b.total_remaining or b.remaining or 0,
            b.security_deposit or 0,
            b.common_notes or "",
            b.staff_names or "",
            b.created_at.strftime("%d-%m-%Y %H:%M") if b.created_at else "",
        ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=bookings_{date.today().isoformat()}.csv"}
    )


@app.route("/admin/export/inventory.csv")
@owner_required
def export_inventory_csv():
    """Download full inventory as CSV."""
    import csv, io
    from flask import Response
    items = ClothingItem.query.order_by(ClothingItem.category, ClothingItem.name).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["SKU", "Name", "Category", "Sub-Category", "Type", "Size", "Color", "Status", "Notes", "Added On"])
    for i in items:
        writer.writerow([
            i.sku, i.name, i.category, i.sub_category or "", i.item_type or "",
            i.size or "", i.color or "", i.status, i.condition_notes or "",
            i.created_at.strftime("%d-%m-%Y") if i.created_at else ""
        ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=inventory_{date.today().isoformat()}.csv"}
    )


# ── Reset All Data ───────────────────────────────────────────────────────────

@app.route("/admin/reset-data", methods=["GET", "POST"])
@owner_required
def reset_all_data():
    if request.method == "POST":
        confirm_text = request.form.get("confirm_text", "").strip()
        what = request.form.get("what", "")

        if confirm_text != "DELETE ALL DATA":
            flash("Confirmation text did not match. Nothing was deleted.", "error")
            return redirect(url_for("reset_all_data"))

        try:
            if what == "bookings_only":
                # Delete all booking-related data only
                BookingItem.query.delete()
                Booking.query.delete()
                Customer.query.delete()
                db.session.commit()
                flash("All bookings and customer records have been deleted.", "success")

            elif what == "everything":
                # Delete ALL operational data; preserve Users + inventory + categories
                BookingItem.query.delete()
                Booking.query.delete()
                Customer.query.delete()
                RentalItem.query.delete()
                Rental.query.delete()
                Payment.query.delete()
                Invoice.query.delete()
                StaffAttendance.query.delete()
                SupplierPurchase.query.delete()
                Supplier.query.delete()
                # Reset clothing item status to available
                ClothingItem.query.update({"status": "available"})
                db.session.commit()
                flash("All data erased. Inventory items reset to available. Users and categories kept.", "success")

            elif what == "full_wipe":
                # Absolute full wipe — everything including inventory and users except THIS owner
                current_user_id = session.get("user_id")
                BookingItem.query.delete()
                Booking.query.delete()
                Customer.query.delete()
                RentalItem.query.delete()
                Rental.query.delete()
                Payment.query.delete()
                Invoice.query.delete()
                StaffAttendance.query.delete()
                SupplierPurchase.query.delete()
                Supplier.query.delete()
                Staff.query.delete()
                ClothingItem.query.delete()
                CustomCategory.query.delete()
                # Delete all users except current
                User.query.filter(User.id != current_user_id).delete()
                db.session.commit()
                flash("Full wipe complete. Only your owner account has been kept.", "success")

            else:
                flash("Unknown reset option.", "error")

        except Exception as e:
            db.session.rollback()
            flash(f"Error during reset: {e}", "error")

        return redirect(url_for("dashboard"))

    return render_template("admin/reset_data.html")


# ── Recycle Bin ──────────────────────────────────────────────────────────────

@app.route("/recycle-bin")
@login_required
def recycle_bin():
    cancelled = Booking.query.filter_by(status="cancelled").order_by(Booking.created_at.desc()).all()
    return render_template("booking/recycle_bin.html", bookings=cancelled)


@app.route("/recycle-bin/<int:id>/restore", methods=["POST"])
@owner_required
def restore_booking(id):
    booking = Booking.query.get_or_404(id)
    if booking.status != "cancelled":
        flash("Only cancelled bookings can be restored.", "error")
        return redirect(url_for("recycle_bin"))
    booking.status = "booked"
    # Re-mark all items as rented
    for bi in booking.booking_items:
        item = ClothingItem.query.get(bi.item_id)
        if item:
            item.status = "rented"
    if booking.item_id and not booking.booking_items:
        item = ClothingItem.query.get(booking.item_id)
        if item:
            item.status = "rented"
    db.session.commit()
    flash(f"Booking Serial #{booking.monthly_serial:02d} restored.", "success")
    return redirect(url_for("recycle_bin"))


@app.route("/recycle-bin/<int:id>/delete", methods=["POST"])
@owner_required
def permanent_delete_booking(id):
    booking = Booking.query.get_or_404(id)
    if booking.status != "cancelled":
        flash("Only cancelled bookings can be permanently deleted.", "error")
        return redirect(url_for("recycle_bin"))
    for bi in booking.booking_items:
        db.session.delete(bi)
    db.session.delete(booking)
    db.session.commit()
    flash(f"Booking permanently deleted.", "success")
    return redirect(url_for("recycle_bin"))


# ── Remaining to Deliver (full page) ─────────────────────────────────────────────

@app.route("/remaining-to-deliver")
@login_required
def remaining_to_deliver_page():
    today = date.today()
    bookings = Booking.query.filter(
        Booking.delivery_date <= today,
        Booking.status == "booked"
    ).order_by(Booking.delivery_date.asc(), Booking.delivery_time).all()
    return render_template("remaining_to_deliver.html", bookings=bookings, today=today)


# ── Free Items Page ──────────────────────────────────────────────────────────────

@app.route("/free-items")
@login_required
def free_items_page():
    return render_template("free_items.html", today=date.today(), categories=CATEGORIES, sizes=SIZES)


# ── Booking List Page ──────────────────────────────────────────────────────────────

@app.route("/booking-list")
@login_required
def booking_list_page():
    return render_template("booking/booking_list.html", today=date.today(), categories=CATEGORIES)


# ── Packing List ─────────────────────────────────────────────────────────────

@app.route("/packing-list")
@login_required
def packing_list_page():
    return render_template("booking/packing_list.html", today=date.today())


@app.route("/api/packing-list")
@login_required
def api_packing_list():
    """Items with status=booked in the given delivery date range – to be packed."""
    delivery_date_str = request.args.get("delivery_date", "")
    return_date_str = request.args.get("return_date", "")
    category_filter = request.args.get("category", "")

    query = Booking.query.filter(Booking.status == "booked")

    try:
        if delivery_date_str and return_date_str:
            d_date = date.fromisoformat(delivery_date_str)
            r_date = date.fromisoformat(return_date_str)
            query = query.filter(Booking.delivery_date >= d_date, Booking.delivery_date <= r_date)
        elif delivery_date_str:
            d_date = date.fromisoformat(delivery_date_str)
            query = query.filter(Booking.delivery_date == d_date)
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400

    bookings = query.order_by(Booking.delivery_date.asc(), Booking.delivery_time).all()

    results = []
    for b in bookings:
        items_data = []
        if b.booking_items:
            for bi in b.booking_items:
                if category_filter and bi.category != category_filter:
                    continue
                item_obj = ClothingItem.query.get(bi.item_id)

                # Check if this item is returning from another customer on b.delivery_date
                ret_warning = None
                returning = Booking.query.filter(
                    Booking.id != b.id,
                    Booking.return_date == b.delivery_date,
                    Booking.status.in_(["booked", "delivered"]),
                ).all()
                for rb in returning:
                    rb_ids = [rbi.item_id for rbi in rb.booking_items] if rb.booking_items else ([rb.item_id] if rb.item_id else [])
                    if bi.item_id in rb_ids:
                        ret_warning = {
                            "customer": rb.customer_name,
                            "serial_no": rb.monthly_serial,
                            "return_time": rb.return_time,
                            "venue": rb.venue or "",
                            "contact": rb.contact_1,
                        }
                        break

                items_data.append({
                    "bi_id": bi.id,
                    "dress_name": bi.dress_name,
                    "display_name": dress_display_name(bi.dress_name, bi.category, bi.size or (item_obj.size if item_obj else "")),
                    "category": bi.category or "",
                    "size": bi.size or (item_obj.size if item_obj else ""),
                    "price": bi.price,
                    "advance": bi.advance,
                    "remaining": bi.remaining,
                    "notes": bi.notes or "",
                    "photo": item_obj.photo if item_obj else "",
                    "prepared_by": bi.prepared_by or "",
                    "checked_by": bi.checked_by or "",
                    "is_packed_ready": bool(bi.is_packed_ready),
                    "packing_note": bi.packing_note or "",
                    "returning_warning": ret_warning,
                })
        elif b.dress_name:
            if not category_filter:
                items_data.append({
                    "bi_id": None,
                    "dress_name": b.dress_name,
                    "category": "",
                    "size": "",
                    "price": b.price,
                    "advance": b.advance,
                    "remaining": b.remaining,
                    "notes": b.notes or "",
                    "photo": "",
                    "prepared_by": "",
                    "checked_by": "",
                    "is_packed_ready": False,
                    "packing_note": "",
                    "returning_warning": None,
                })
        if not items_data:
            continue
        results.append({
            "id": b.id,
            "serial_no": b.monthly_serial,
            "customer_name": b.customer_name,
            "customer_address": b.customer_address,
            "contact_1": b.contact_1,
            "whatsapp_no": b.whatsapp_no or "",
            "venue": b.venue or "",
            "delivery_date": b.delivery_date.isoformat(),
            "delivery_time": b.delivery_time,
            "return_date": b.return_date.isoformat(),
            "return_time": b.return_time,
            "total_price": b.total_price or b.price,
            "total_advance": b.total_advance or b.advance,
            "total_remaining": b.total_remaining or b.remaining,
            "security_deposit": b.security_deposit,
            "common_notes": b.common_notes or "",
            "staff_names": b.staff_names or "",
            "items": items_data,
        })
    return jsonify(results)


@app.route("/api/packing-list/save-item", methods=["POST"])
@login_required
def api_packing_save_item():
    """Save packing fields for a single BookingItem."""
    data = request.get_json(force=True)
    bi_id = data.get("bi_id")

    if not bi_id:
        return jsonify({"ok": False, "error": "Missing bi_id"}), 400

    bi = BookingItem.query.get(bi_id)
    if not bi:
        return jsonify({"ok": False, "error": "Item not found"}), 404

    if "prepared_by" in data:
        bi.prepared_by = (data["prepared_by"] or "").strip() or None
    if "checked_by" in data:
        bi.checked_by = (data["checked_by"] or "").strip() or None
    if "is_packed_ready" in data:
        bi.is_packed_ready = bool(data["is_packed_ready"])
    if "packing_note" in data:
        bi.packing_note = (data["packing_note"] or "").strip() or None

    db.session.commit()
    return jsonify({
        "ok": True,
        "prepared_by": bi.prepared_by or "",
        "checked_by": bi.checked_by or "",
        "is_packed_ready": bool(bi.is_packed_ready),
        "packing_note": bi.packing_note or "",
    })


@app.route("/api/booking-list")
@login_required
def api_booking_list():
    """Bookings overlapping a date range + dresses unavailable (delivered before from-date, still out)."""
    delivery_date_str = request.args.get("delivery_date", "")
    return_date_str = request.args.get("return_date", "")
    category_filter = request.args.get("category", "")
    delivery_time = request.args.get("delivery_time", "")
    return_time = request.args.get("return_time", "")

    if not delivery_date_str:
        return jsonify({"bookings": [], "unavailable": []})

    try:
        d_date = date.fromisoformat(delivery_date_str)
        r_date = date.fromisoformat(return_date_str) if return_date_str else d_date
    except ValueError:
        return jsonify({"error": "Invalid date"}), 400

    if r_date < d_date:
        r_date = d_date

    def _build_items_for_booking(b, for_category_only=True):
        items_data = []
        if b.booking_items:
            for bi in b.booking_items:
                if for_category_only and category_filter and bi.category != category_filter:
                    continue
                item_obj = ClothingItem.query.get(bi.item_id)

                ret_warning = None
                ret_bookings = Booking.query.filter(
                    Booking.id != b.id,
                    Booking.status.in_(["booked", "delivered"]),
                    Booking.return_date == b.delivery_date,
                ).all()
                for rb in ret_bookings:
                    r_ids = [rbi.item_id for rbi in rb.booking_items] if rb.booking_items else ([rb.item_id] if rb.item_id else [])
                    if bi.item_id in r_ids:
                        ret_warning = {
                            "customer": rb.customer_name,
                            "total_rent": rb.total_price or rb.price,
                            "delivery_time": rb.return_time,
                            "serial_no": rb.monthly_serial,
                            "venue": rb.venue or "",
                        }
                        break

                del_warning = None
                del_bookings = Booking.query.filter(
                    Booking.id != b.id,
                    Booking.status.in_(["booked", "delivered"]),
                    Booking.delivery_date == b.return_date,
                ).all()
                for db_item in del_bookings:
                    d_ids = [dbi.item_id for dbi in db_item.booking_items] if db_item.booking_items else ([db_item.item_id] if db_item.item_id else [])
                    if bi.item_id in d_ids:
                        del_warning = {
                            "customer": db_item.customer_name,
                            "total_rent": db_item.total_price or db_item.price,
                            "delivery_time": db_item.delivery_time,
                            "serial_no": db_item.monthly_serial,
                            "venue": db_item.venue or "",
                        }
                        break

                sz = bi.size or (item_obj.size if item_obj else "")
                items_data.append({
                    "dress_name": bi.dress_name,
                    "display_name": dress_display_name(bi.dress_name, bi.category, sz),
                    "category": bi.category,
                    "size": sz,
                    "price": bi.price,
                    "notes": bi.notes or "",
                    "photo": item_obj.photo or "" if item_obj else "",
                    "returning_warning": ret_warning,
                    "booked_warning": del_warning,
                })
        elif b.item_id:
            item_obj = ClothingItem.query.get(b.item_id)
            if for_category_only and category_filter and item_obj and item_obj.category != category_filter:
                return []
            items_data.append({
                "dress_name": b.dress_name,
                "display_name": dress_display_name(b.dress_name, item_obj.category if item_obj else "", item_obj.size if item_obj else ""),
                "category": item_obj.category if item_obj else "",
                "size": item_obj.size if item_obj else "",
                "price": b.price,
                "notes": b.notes or "",
                "photo": item_obj.photo or "" if item_obj else "",
                "returning_warning": None,
                "booked_warning": None,
            })
        return items_data

    def _serialize_booking(b):
        items_data = _build_items_for_booking(b)
        if not items_data and category_filter:
            return None
        return {
            "id": b.id,
            "booking_number": b.booking_number,
            "serial_no": b.monthly_serial,
            "customer_name": b.customer_name,
            "contact_1": b.contact_1 or "",
            "venue": b.venue or "",
            "delivery_date": b.delivery_date.strftime("%d %b %Y"),
            "delivery_date_iso": b.delivery_date.isoformat(),
            "delivery_time": b.delivery_time,
            "return_date": b.return_date.strftime("%d %b %Y"),
            "return_date_iso": b.return_date.isoformat(),
            "return_time": b.return_time,
            "total_price": b.total_price or b.price,
            "total_advance": b.total_advance or b.advance,
            "total_remaining": b.total_remaining or b.remaining,
            "common_notes": b.common_notes or "",
            "status": b.status,
            "items": items_data,
        }

    # Main list: deliveries scheduled between From and To (both dates inclusive)
    main_query = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.delivery_date >= d_date,
        Booking.delivery_date <= r_date,
    )
    if delivery_time:
        main_query = main_query.filter(Booking.delivery_time == delivery_time)
    if return_time:
        main_query = main_query.filter(Booking.return_time == return_time)

    main_bookings = main_query.order_by(
        Booking.delivery_date.asc(), Booking.delivery_time
    ).all()

    results = []
    for b in main_bookings:
        row = _serialize_booking(b)
        if row:
            results.append(row)

    # Unavailable: delivered before From, return before To, still out when period starts
    unavail_query = Booking.query.filter(
        Booking.status.in_(["booked", "delivered"]),
        Booking.delivery_date < d_date,
        Booking.return_date >= d_date,
        Booking.return_date < r_date,
    )
    if delivery_time:
        unavail_query = unavail_query.filter(Booking.delivery_time == delivery_time)
    if return_time:
        unavail_query = unavail_query.filter(Booking.return_time == return_time)

    unavail_bookings = unavail_query.order_by(
        Booking.delivery_date.asc(), Booking.return_time
    ).all()

    unavailable = []
    for b in unavail_bookings:
        row = _serialize_booking(b)
        if row:
            row["reason"] = (
                f"Delivered {b.delivery_date.strftime('%d %b %Y')} (before {d_date.strftime('%d %b %Y')}) "
                f"— returns {b.return_date.strftime('%d %b %Y')} (before {r_date.strftime('%d %b %Y')})"
            )
            unavailable.append(row)

    return jsonify({
        "bookings": results,
        "unavailable": unavailable,
        "from_date": d_date.strftime("%d %b %Y"),
        "to_date": r_date.strftime("%d %b %Y"),
    })


# ── Finance: Daily Sale ──────────────────────────────────────────────────────

@app.route("/finance/daily-sale")
@owner_required
def finance_daily_sale():
    return render_template("finance/daily_sale.html", today=date.today())


@app.route("/api/finance/daily-sale")
@login_required
def api_finance_daily_sale():
    """Advance collected + remaining balance collected on a given date, category-wise."""
    target_date_str = request.args.get("date", date.today().isoformat())
    try:
        target_date = date.fromisoformat(target_date_str)
    except ValueError:
        target_date = date.today()

    # Advance collected: bookings created on target_date
    bookings_today = Booking.query.filter(
        db.func.date(Booking.created_at) == target_date,
        Booking.status != "cancelled",
    ).all()

    # Remaining balance collected: bookings delivered on target_date (remaining paid on delivery)
    delivered_today = Booking.query.filter(
        Booking.delivery_date == target_date,
        Booking.status.in_(["delivered", "returned"]),
    ).all()

    advance_by_cat = {}
    remaining_by_cat = {}
    advance_mens = 0
    advance_womens = 0
    advance_jewellery = 0
    remaining_mens = 0
    remaining_womens = 0
    remaining_jewellery = 0

    for b in bookings_today:
        if b.booking_items:
            for bi in b.booking_items:
                cat = bi.category or "Other"
                advance_by_cat[cat] = advance_by_cat.get(cat, 0) + bi.advance
                if cat in MENS_CATEGORIES:
                    advance_mens += bi.advance
                elif cat in WOMENS_CATEGORIES:
                    advance_womens += bi.advance
                elif cat in JEWELLERY_CATEGORIES:
                    advance_jewellery += bi.advance
        else:
            advance_by_cat["Other"] = advance_by_cat.get("Other", 0) + (b.total_advance or b.advance)

    for b in delivered_today:
        remaining_amt = b.remaining_collected or 0
        if b.booking_items and remaining_amt > 0:
            for bi in b.booking_items:
                cat = bi.category or "Other"
                share = remaining_amt * (bi.remaining / b.total_remaining) if b.total_remaining else 0
                remaining_by_cat[cat] = remaining_by_cat.get(cat, 0) + share
                if cat in MENS_CATEGORIES:
                    remaining_mens += share
                elif cat in WOMENS_CATEGORIES:
                    remaining_womens += share
                elif cat in JEWELLERY_CATEGORIES:
                    remaining_jewellery += share
        elif not b.booking_items:
            remaining_by_cat["Other"] = remaining_by_cat.get("Other", 0) + remaining_amt

    total_advance = sum(advance_by_cat.values())
    total_remaining = sum(remaining_by_cat.values())
    total_sale = total_advance + total_remaining

    return jsonify({
        "date": target_date.strftime("%d %b %Y"),
        "advance_by_category": advance_by_cat,
        "remaining_by_category": remaining_by_cat,
        "total_advance": total_advance,
        "total_remaining_collected": total_remaining,
        "total_sale": total_sale,
        "advance_mens": advance_mens,
        "advance_womens": advance_womens,
        "advance_jewellery": advance_jewellery,
        "remaining_mens": remaining_mens,
        "remaining_womens": remaining_womens,
        "remaining_jewellery": remaining_jewellery,
    })


# ── Finance: Daily Booking Amount ────────────────────────────────────────────

@app.route("/finance/daily-booking")
@owner_required
def finance_daily_booking():
    return render_template("finance/daily_booking.html", today=date.today())


@app.route("/api/finance/daily-booking")
@login_required
def api_finance_daily_booking():
    """Total booking amounts for bookings created on a given date."""
    target_date_str = request.args.get("date", date.today().isoformat())
    try:
        target_date = date.fromisoformat(target_date_str)
    except ValueError:
        target_date = date.today()

    bookings = Booking.query.filter(
        db.func.date(Booking.created_at) == target_date,
        Booking.status != "cancelled",
    ).all()

    total_by_cat = {}
    mens_total = 0
    womens_total = 0
    jewellery_total = 0
    for b in bookings:
        if b.booking_items:
            for bi in b.booking_items:
                cat = bi.category or "Other"
                total_by_cat[cat] = total_by_cat.get(cat, 0) + bi.price
                if cat in MENS_CATEGORIES:
                    mens_total += bi.price
                elif cat in WOMENS_CATEGORIES:
                    womens_total += bi.price
                elif cat in JEWELLERY_CATEGORIES:
                    jewellery_total += bi.price
        else:
            total_by_cat["Other"] = total_by_cat.get("Other", 0) + (b.total_price or b.price)

    return jsonify({
        "date": target_date.strftime("%d %b %Y"),
        "total_by_category": total_by_cat,
        "grand_total": sum(total_by_cat.values()),
        "mens_total": mens_total,
        "womens_total": womens_total,
        "jewellery_total": jewellery_total,
    })


# ── Finance: Monthly Sale ────────────────────────────────────────────────────

@app.route("/finance/monthly-sale")
@owner_required
def finance_monthly_sale():
    return render_template("finance/monthly_sale.html", today=date.today())


@app.route("/api/finance/monthly-sale")
@login_required
def api_finance_monthly_sale():
    target_month = request.args.get("month", date.today().strftime("%Y-%m"))
    try:
        year, month = map(int, target_month.split("-"))
    except ValueError:
        year, month = date.today().year, date.today().month

    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)

    bookings = Booking.query.filter(
        db.func.date(Booking.created_at) >= month_start,
        db.func.date(Booking.created_at) < month_end,
        Booking.status != "cancelled",
    ).all()

    total_advance = sum(b.total_advance or b.advance for b in bookings)
    total_remaining = sum(b.total_remaining or b.remaining for b in bookings)
    total_sale = total_advance + total_remaining

    mens_total = 0
    womens_total = 0
    jewellery_total = 0
    for b in bookings:
        if b.booking_items:
            for bi in b.booking_items:
                if bi.category in MENS_CATEGORIES:
                    mens_total += bi.price
                elif bi.category in WOMENS_CATEGORIES:
                    womens_total += bi.price
                elif bi.category in JEWELLERY_CATEGORIES:
                    jewellery_total += bi.price

    return jsonify({
        "month": f"{year}-{month:02d}",
        "total_advance": total_advance,
        "total_remaining": total_remaining,
        "total_sale": total_sale,
        "booking_count": len(bookings),
        "mens_total": mens_total,
        "womens_total": womens_total,
        "jewellery_total": jewellery_total,
    })


# ── Finance: Yearly Sale ─────────────────────────────────────────────────────

@app.route("/finance/yearly-sale")
@owner_required
def finance_yearly_sale():
    return render_template("finance/yearly_sale.html", today=date.today())


@app.route("/api/finance/yearly-sale")
@login_required
def api_finance_yearly_sale():
    """Financial year: April to next March. Also supports custom date range."""
    from_str = request.args.get("from", "")
    to_str = request.args.get("to", "")

    if from_str and to_str:
        try:
            from_date = date.fromisoformat(from_str)
            to_date = date.fromisoformat(to_str)
        except ValueError:
            from_date = date(date.today().year, 4, 1)
            to_date = date(date.today().year + 1, 3, 31)
    else:
        today = date.today()
        if today.month >= 4:
            from_date = date(today.year, 4, 1)
            to_date = date(today.year + 1, 3, 31)
        else:
            from_date = date(today.year - 1, 4, 1)
            to_date = date(today.year, 3, 31)

    bookings = Booking.query.filter(
        db.func.date(Booking.created_at) >= from_date,
        db.func.date(Booking.created_at) <= to_date,
        Booking.status != "cancelled",
    ).all()

    total_advance = sum(b.total_advance or b.advance for b in bookings)
    total_remaining = sum(b.total_remaining or b.remaining for b in bookings)
    total_sale = total_advance + total_remaining

    # Monthly breakdown
    monthly = {}
    cat_totals = {}
    mens_total = 0
    womens_total = 0
    jewellery_total = 0
    for b in bookings:
        m_key = b.created_at.strftime("%Y-%m") if b.created_at else "Unknown"
        monthly[m_key] = monthly.get(m_key, 0) + (b.total_price or b.price)
        if b.booking_items:
            for bi in b.booking_items:
                cat = bi.category or "Other"
                cat_totals[cat] = cat_totals.get(cat, 0) + bi.price
                if cat in MENS_CATEGORIES:
                    mens_total += bi.price
                elif cat in WOMENS_CATEGORIES:
                    womens_total += bi.price
                elif cat in JEWELLERY_CATEGORIES:
                    jewellery_total += bi.price

    return jsonify({
        "from": from_date.isoformat(),
        "to": to_date.isoformat(),
        "total_advance": total_advance,
        "total_remaining": total_remaining,
        "total_sale": total_sale,
        "monthly_breakdown": monthly,
        "category_totals": cat_totals,
        "mens_total": mens_total,
        "womens_total": womens_total,
        "jewellery_total": jewellery_total,
        "booking_count": len(bookings),
    })


# ── Finance: Top Performer ───────────────────────────────────────────────────

@app.route("/finance/top-performer")
@owner_required
def finance_top_performer():
    return render_template("finance/top_performer.html", today=date.today())


@app.route("/api/finance/top-performer")
@login_required
def api_finance_top_performer():
    from_str = request.args.get("from", "")
    to_str = request.args.get("to", "")
    category_filter = request.args.get("category", "")

    if from_str and to_str:
        try:
            from_date = date.fromisoformat(from_str)
            to_date = date.fromisoformat(to_str)
        except ValueError:
            from_date = date.today().replace(day=1)
            to_date = date.today()
    else:
        from_date = date.today().replace(day=1)
        to_date = date.today()

    bookings = Booking.query.filter(
        db.func.date(Booking.created_at) >= from_date,
        db.func.date(Booking.created_at) <= to_date,
        Booking.status != "cancelled",
    ).all()

    product_stats = {}
    for b in bookings:
        if b.booking_items:
            for bi in b.booking_items:
                if category_filter and bi.category != category_filter:
                    continue
                item_obj = ClothingItem.query.get(bi.item_id)
                key = f"{bi.item_id}"
                if key not in product_stats:
                    product_stats[key] = {
                        "name": bi.dress_name,
                        "category": bi.category,
                        "size": item_obj.size if item_obj else "",
                        "photo": item_obj.photo if item_obj else "",
                        "bookings": 0,
                        "total_earned": 0,
                    }
                product_stats[key]["bookings"] += 1
                product_stats[key]["total_earned"] += bi.price

    ranked = sorted(product_stats.values(), key=lambda x: x["total_earned"], reverse=True)
    return jsonify(ranked)


# ── Finance: Category Analysis ───────────────────────────────────────────────

def _normalize_whatsapp_number(phone):
    """Return 91XXXXXXXXXX for WhatsApp bulk tools."""
    if not phone:
        return ""
    digits = "".join(c for c in str(phone) if c.isdigit())
    if len(digits) == 10:
        return "91" + digits
    if len(digits) == 12 and digits.startswith("91"):
        return digits
    if len(digits) == 11 and digits.startswith("0"):
        return "91" + digits[1:]
    return digits


def _get_customers_by_category(category=None):
    """Build unique customer contacts from bookings, optionally filtered by dress category."""
    seen = {}
    bookings = Booking.query.filter(Booking.status != "cancelled").all()
    for b in bookings:
        cats_for_booking = set()
        if b.booking_items:
            for bi in b.booking_items:
                if bi.category:
                    cats_for_booking.add(bi.category)
        elif b.item_id:
            item = ClothingItem.query.get(b.item_id)
            if item and item.category:
                cats_for_booking.add(item.category)

        if category and category not in cats_for_booking:
            continue

        phone = (b.whatsapp_no or b.contact_1 or "").strip()
        if not phone:
            continue
        key = _normalize_whatsapp_number(phone) or phone
        if key not in seen:
            seen[key] = {
                "name": b.customer_name,
                "phone": b.contact_1 or phone,
                "whatsapp": b.whatsapp_no or b.contact_1 or phone,
                "whatsapp_number": _normalize_whatsapp_number(b.whatsapp_no or b.contact_1),
                "address": b.customer_address or "",
                "categories": set(cats_for_booking),
            }
        else:
            seen[key]["categories"].update(cats_for_booking)

    # Merge Customer table records
    for c in Customer.query.all():
        phone = (c.phone or "").strip()
        if not phone:
            continue
        key = _normalize_whatsapp_number(phone) or phone
        cust_cats = set()
        if category:
            # Only include if they have bookings in this category
            if key not in seen:
                continue
            cust_cats = seen[key]["categories"]
        if key in seen:
            if c.address and not seen[key]["address"]:
                seen[key]["address"] = c.address
            if c.name:
                seen[key]["name"] = c.name
        elif not category:
            seen[key] = {
                "name": c.name,
                "phone": c.phone,
                "whatsapp": c.phone,
                "whatsapp_number": _normalize_whatsapp_number(c.phone),
                "address": c.address or "",
                "categories": cust_cats,
            }

    result = []
    for v in seen.values():
        cat_label = category or ", ".join(sorted(v["categories"])) or "General"
        if category or v["categories"] or True:
            result.append({
                "name": v["name"],
                "phone": v["phone"],
                "whatsapp": v["whatsapp"],
                "whatsapp_number": v["whatsapp_number"],
                "address": v["address"],
                "category": cat_label,
                "categories": sorted(v["categories"]),
            })
    result.sort(key=lambda x: x["name"].lower())
    return result


def _build_category_analysis(from_date, to_date):
    """Aggregate purchase, stock, and sale data by category for a date range."""
    # ── Purchases from suppliers ──
    purchases = SupplierPurchase.query.filter(
        SupplierPurchase.date >= from_date,
        SupplierPurchase.date <= to_date,
    ).all()

    purchase_by_cat = {}
    returns_by_cat = {}
    gst_by_cat = {}

    for p in purchases:
        cat = (p.category or "Uncategorized").strip() or "Uncategorized"
        if p.transaction_type == "return":
            returns_by_cat[cat] = returns_by_cat.get(cat, 0) + (p.amount or 0)
        else:
            purchase_by_cat[cat] = purchase_by_cat.get(cat, 0) + (p.amount or 0)
            gst_by_cat[cat] = gst_by_cat.get(cat, 0) + (p.gst_amount or 0)

    net_purchase_by_cat = {}
    all_purchase_cats = set(purchase_by_cat) | set(returns_by_cat)
    for cat in all_purchase_cats:
        net_purchase_by_cat[cat] = purchase_by_cat.get(cat, 0) - returns_by_cat.get(cat, 0)

    # ── Current stock count by category ──
    stock_by_cat = {}
    for item in ClothingItem.query.all():
        cat = (item.category or "Uncategorized").strip() or "Uncategorized"
        stock_by_cat[cat] = stock_by_cat.get(cat, 0) + 1

    # ── Sales: advance (booking created in range) + remaining (collected on delivery in range) ──
    advance_by_cat = {}
    remaining_by_cat = {}

    bookings_created = Booking.query.filter(
        Booking.status != "cancelled",
        db.func.date(Booking.created_at) >= from_date,
        db.func.date(Booking.created_at) <= to_date,
    ).all()

    for b in bookings_created:
        if b.booking_items:
            for bi in b.booking_items:
                cat = (bi.category or "Uncategorized").strip() or "Uncategorized"
                advance_by_cat[cat] = advance_by_cat.get(cat, 0) + (bi.advance or 0)
        else:
            cat = "Uncategorized"
            advance_by_cat[cat] = advance_by_cat.get(cat, 0) + (b.total_advance or b.advance or 0)

    delivered_in_range = Booking.query.filter(
        Booking.delivery_date >= from_date,
        Booking.delivery_date <= to_date,
        Booking.status.in_(["delivered", "returned"]),
    ).all()

    for b in delivered_in_range:
        remaining_amt = b.remaining_collected or 0
        if remaining_amt <= 0:
            continue
        if b.booking_items and b.total_remaining:
            for bi in b.booking_items:
                cat = (bi.category or "Uncategorized").strip() or "Uncategorized"
                share = remaining_amt * ((bi.remaining or 0) / b.total_remaining)
                remaining_by_cat[cat] = remaining_by_cat.get(cat, 0) + share
        elif b.booking_items:
            per_item = remaining_amt / len(b.booking_items)
            for bi in b.booking_items:
                cat = (bi.category or "Uncategorized").strip() or "Uncategorized"
                remaining_by_cat[cat] = remaining_by_cat.get(cat, 0) + per_item
        else:
            cat = "Uncategorized"
            remaining_by_cat[cat] = remaining_by_cat.get(cat, 0) + remaining_amt

    sale_by_cat = {}
    all_sale_cats = set(advance_by_cat) | set(remaining_by_cat)
    for cat in all_sale_cats:
        sale_by_cat[cat] = round(advance_by_cat.get(cat, 0) + remaining_by_cat.get(cat, 0), 2)

    # All categories union
    all_cats = sorted(
        set(net_purchase_by_cat) | set(stock_by_cat) | set(sale_by_cat) | set(gst_by_cat),
        key=lambda c: c.lower(),
    )

    def _pie_data(amounts_dict):
        total = sum(amounts_dict.values())
        items = []
        for cat, amt in sorted(amounts_dict.items(), key=lambda x: -x[1]):
            if amt <= 0:
                continue
            pct = round(amt / total * 100, 1) if total else 0
            items.append({"category": cat, "amount": round(amt, 2), "percent": pct})
        return items, round(total, 2)

    purchase_pie, total_purchase = _pie_data({k: v for k, v in net_purchase_by_cat.items() if v > 0})
    sale_pie, total_sale = _pie_data(sale_by_cat)

    compare = []
    compare_cats = sorted(set(net_purchase_by_cat) | set(sale_by_cat), key=lambda c: c.lower())
    for cat in compare_cats:
        pur = round(net_purchase_by_cat.get(cat, 0), 2)
        sal = round(sale_by_cat.get(cat, 0), 2)
        compare.append({
            "category": cat,
            "purchase": pur,
            "sale": sal,
            "profit": round(sal - pur, 2),
            "stock": stock_by_cat.get(cat, 0),
        })

    rows = []
    for cat in all_cats:
        rows.append({
            "category": cat,
            "purchase": round(net_purchase_by_cat.get(cat, 0), 2),
            "purchase_gross": round(purchase_by_cat.get(cat, 0), 2),
            "returns": round(returns_by_cat.get(cat, 0), 2),
            "gst": round(gst_by_cat.get(cat, 0), 2),
            "stock": stock_by_cat.get(cat, 0),
            "advance": round(advance_by_cat.get(cat, 0), 2),
            "remaining": round(remaining_by_cat.get(cat, 0), 2),
            "sale": round(sale_by_cat.get(cat, 0), 2),
        })

    return {
        "from": from_date.strftime("%d %b %Y"),
        "to": to_date.strftime("%d %b %Y"),
        "total_purchase": total_purchase,
        "total_sale": total_sale,
        "total_gst": round(sum(gst_by_cat.values()), 2),
        "total_stock": sum(stock_by_cat.values()),
        "purchase_pie": purchase_pie,
        "sale_pie": sale_pie,
        "compare": compare,
        "rows": rows,
    }


@app.route("/finance/category-analysis")
@owner_required
def finance_category_analysis():
    today = date.today()
    month_start = today.replace(day=1)
    return render_template(
        "finance/category_analysis.html",
        today=today,
        month_start=month_start,
    )


@app.route("/api/finance/category-analysis")
@owner_required
def api_finance_category_analysis():
    from_str = request.args.get("from", "")
    to_str = request.args.get("to", "")
    try:
        from_date = date.fromisoformat(from_str) if from_str else date.today().replace(day=1)
        to_date = date.fromisoformat(to_str) if to_str else date.today()
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400
    if to_date < from_date:
        return jsonify({"error": "End date must be on or after start date"}), 400
    return jsonify(_build_category_analysis(from_date, to_date))


# ── Finance: Suppliers ───────────────────────────────────────────────────────

@app.route("/finance/suppliers")
@owner_required
def finance_suppliers():
    suppliers = Supplier.query.order_by(Supplier.name).all()
    return render_template("finance/suppliers.html", suppliers=suppliers, today=date.today())


@app.route("/finance/suppliers/add", methods=["POST"])
@owner_required
def add_supplier():
    name = request.form["name"].strip()
    phone = request.form.get("phone", "").strip()
    address = request.form.get("address", "").strip()
    s = Supplier(name=name, phone=phone, address=address)
    db.session.add(s)
    db.session.commit()
    flash(f"Supplier '{name}' added.", "success")
    return redirect(url_for("finance_suppliers"))


@app.route("/finance/suppliers/<int:id>/purchase", methods=["POST"])
@owner_required
def add_purchase(id):
    supplier = Supplier.query.get_or_404(id)
    amount = float(request.form.get("amount", 0) or 0)
    gst_amount = float(request.form.get("gst_amount", 0) or 0)
    p = SupplierPurchase(
        supplier_id=id,
        item_description=request.form["item_description"].strip(),
        category=request.form.get("category", ""),
        amount=amount,
        gst_amount=gst_amount,
        transaction_type="purchase",
        date=date.fromisoformat(request.form["date"]) if request.form.get("date") else date.today(),
        notes=request.form.get("notes", ""),
    )
    db.session.add(p)
    db.session.commit()
    flash(f"Purchase ₹{amount:,.0f} (GST ₹{gst_amount:,.0f}) recorded for '{supplier.name}'.", "success")
    return redirect(url_for("finance_suppliers"))


@app.route("/finance/suppliers/<int:id>/return", methods=["POST"])
@owner_required
def add_supplier_return(id):
    """Record goods returned to a supplier — deducts from the supplier's total."""
    supplier = Supplier.query.get_or_404(id)
    amount = float(request.form.get("amount", 0) or 0)
    p = SupplierPurchase(
        supplier_id=id,
        item_description=request.form["item_description"].strip(),
        category=request.form.get("category", ""),
        amount=amount,
        gst_amount=0,
        transaction_type="return",
        date=date.fromisoformat(request.form["date"]) if request.form.get("date") else date.today(),
        notes=request.form.get("notes", ""),
    )
    db.session.add(p)
    db.session.commit()
    flash(f"Return of ₹{amount:,.0f} recorded for '{supplier.name}'. Amount deducted from total.", "success")
    return redirect(url_for("finance_suppliers"))


# ── Staff Work ───────────────────────────────────────────────────────────────

@app.route("/staff-work")
@owner_required
def staff_work_page():
    return render_template("staff/staff_work.html", today=date.today())


@app.route("/api/staff-work")
def api_staff_work():
    from_str = request.args.get("from", date.today().replace(day=1).isoformat())
    to_str = request.args.get("to", date.today().isoformat())
    try:
        from_date = date.fromisoformat(from_str)
        to_date = date.fromisoformat(to_str)
    except ValueError:
        from_date = date.today().replace(day=1)
        to_date = date.today()

    bookings = Booking.query.filter(
        db.func.date(Booking.created_at) >= from_date,
        db.func.date(Booking.created_at) <= to_date,
        Booking.status != "cancelled",
    ).all()

    staff_stats = {}
    for b in bookings:
        if not b.staff_names:
            continue
        names = [n.strip() for n in b.staff_names.split(",") if n.strip()]
        total = b.total_price or b.price
        split_amount = total / len(names) if names else 0
        item_count = len(b.booking_items) if b.booking_items else 1
        for name in names:
            if name not in staff_stats:
                staff_stats[name] = {"name": name, "bookings": 0, "dresses": 0, "amount": 0}
            staff_stats[name]["bookings"] += 1
            staff_stats[name]["dresses"] += item_count
            staff_stats[name]["amount"] += split_amount

    return jsonify(list(staff_stats.values()))


# ── Staff Attendance ─────────────────────────────────────────────────────────

@app.route("/staff-attendance")
@login_required
def staff_attendance_page():
    staff_list = Staff.query.filter_by(active=True).order_by(Staff.name).all()
    all_users = User.query.order_by(User.username).all()
    return render_template("staff/attendance.html", staff_list=staff_list, today=date.today(), all_users=all_users)


@app.route("/staff/add", methods=["POST"])
@owner_required
def add_staff():
    name = request.form["name"].strip()
    phone = request.form.get("phone", "").strip()
    role = request.form.get("role", "staff")
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()

    s = Staff(name=name, phone=phone)
    db.session.add(s)
    db.session.flush()  # get s.id

    # Create login account if username provided
    if username and password:
        if User.query.filter_by(username=username).first():
            db.session.rollback()
            flash(f"Username '{username}' is already taken.", "error")
            return redirect(url_for("staff_attendance_page"))
        u = User(username=username, role=role, staff_id=s.id)
        u.set_password(password)
        db.session.add(u)

    db.session.commit()
    flash(f"Staff '{name}' added." + (f" Login account '{username}' created." if username else ""), "success")
    return redirect(url_for("staff_attendance_page"))


@app.route("/staff/<int:id>/remove", methods=["POST"])
@owner_required
def remove_staff(id):
    s = Staff.query.get_or_404(id)
    s.active = False
    # Deactivate linked user too
    linked_user = User.query.filter_by(staff_id=id).first()
    if linked_user:
        linked_user.active = False
    db.session.commit()
    flash(f"Staff '{s.name}' removed.", "success")
    return redirect(url_for("staff_attendance_page"))


@app.route("/staff/attendance/save", methods=["POST"])
@login_required
def save_attendance():
    att_date = date.fromisoformat(request.form["date"])
    staff_list = Staff.query.filter_by(active=True).all()
    for s in staff_list:
        status = request.form.get(f"status_{s.id}", "present")
        existing = StaffAttendance.query.filter_by(staff_id=s.id, date=att_date).first()
        if existing:
            existing.status = status
        else:
            db.session.add(StaffAttendance(staff_id=s.id, date=att_date, status=status))
    db.session.commit()
    flash(f"Attendance saved for {att_date.strftime('%d %b %Y')}.", "success")
    return redirect(url_for("staff_attendance_page"))


@app.route("/api/staff/attendance")
def api_staff_attendance():
    month_str = request.args.get("month", date.today().strftime("%Y-%m"))
    try:
        year, month = map(int, month_str.split("-"))
    except ValueError:
        year, month = date.today().year, date.today().month

    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)

    staff_list = Staff.query.filter_by(active=True).order_by(Staff.name).all()
    result = []
    for s in staff_list:
        records = StaffAttendance.query.filter(
            StaffAttendance.staff_id == s.id,
            StaffAttendance.date >= month_start,
            StaffAttendance.date < month_end,
        ).all()
        present = sum(1 for r in records if r.status == "present")
        absent = sum(1 for r in records if r.status == "absent")
        half_day = sum(1 for r in records if r.status == "half_day")
        result.append({
            "id": s.id, "name": s.name,
            "present": present, "absent": absent, "half_day": half_day,
        })
    return jsonify(result)


@app.route("/staff/shop-closed", methods=["POST"])
@owner_required
def mark_shop_closed():
    att_date = date.fromisoformat(request.form["date"])
    staff_list = Staff.query.filter_by(active=True).all()
    for s in staff_list:
        existing = StaffAttendance.query.filter_by(staff_id=s.id, date=att_date).first()
        if existing:
            existing.status = "shop_closed"
        else:
            db.session.add(StaffAttendance(staff_id=s.id, date=att_date, status="shop_closed"))
    db.session.commit()
    flash(f"Shop marked CLOSED for {att_date.strftime('%d %b %Y')}.", "info")
    return redirect(url_for("staff_attendance_page"))


@app.route("/api/staff/attendance-calendar")
def api_staff_attendance_calendar():
    staff_id = request.args.get("staff_id", type=int)
    month_str = request.args.get("month", date.today().strftime("%Y-%m"))
    try:
        year, month = map(int, month_str.split("-"))
    except ValueError:
        year, month = date.today().year, date.today().month

    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)

    records = StaffAttendance.query.filter(
        StaffAttendance.staff_id == staff_id,
        StaffAttendance.date >= month_start,
        StaffAttendance.date < month_end,
    ).all()

    days = {}
    for r in records:
        days[r.date.isoformat()] = r.status
    return jsonify({"days": days})


# ────────────────── SEARCH BOOKING (month-based) ──────────────────
@app.route("/search-booking")
@login_required
def search_booking_page():
    return render_template("booking/search_booking.html")


@app.route("/api/search-booking")
def api_search_booking():
    """Month-based search - same logic as delivery search but for all statuses except cancelled."""
    search_date = request.args.get("date", date.today().isoformat())
    query_text = request.args.get("q", "").strip()

    try:
        ref_date = datetime.strptime(search_date, "%Y-%m-%d").date()
    except ValueError:
        ref_date = date.today()

    year = ref_date.year
    month = ref_date.month

    def get_bookings(y, m):
        m_start = date(y, m, 1)
        if m == 12:
            m_end = date(y + 1, 1, 1)
        else:
            m_end = date(y, m + 1, 1)
        q = Booking.query.filter(
            Booking.delivery_date >= m_start,
            Booking.delivery_date < m_end,
            Booking.status.in_(["booked", "delivered"]),  # exclude cancelled and returned
        )
        if query_text:
            filt = booking_search_filter(query_text)
            if filt is not None:
                q = q.filter(filt)
        return q.order_by(Booking.delivery_date).all()

    results = get_bookings(year, month)
    if not results and query_text:
        prev_m = month - 1 if month > 1 else 12
        prev_y = year if month > 1 else year - 1
        next_m = month + 1 if month < 12 else 1
        next_y = year if month < 12 else year + 1
        results = get_bookings(prev_y, prev_m) + get_bookings(next_y, next_m)

    data = []
    for b in results:
        items_list = booking_items_for_api(b)
        data.append({
            "id": b.id,
            "booking_number": b.booking_number,
            "serial": b.monthly_serial,
            "customer_name": b.customer_name,
            "contact_1": b.contact_1,
            "delivery_date": b.delivery_date.isoformat(),
            "delivery_time": b.delivery_time,
            "return_date": b.return_date.isoformat(),
            "status": b.status,
            "total_price": b.total_price,
            "venue": b.venue or "",
            "items": items_list,
        })
    return jsonify(data)


# ────────────────── ALL RECORD SEARCH (universal) ──────────────────
@app.route("/all-record-search")
@login_required
def all_record_search_page():
    return render_template("booking/all_record_search.html")


@app.route("/api/all-record-search")
def api_all_record_search():
    """Universal search across all months/years including delivered and returned."""
    search_date = request.args.get("date", "")
    query_text = request.args.get("q", "").strip()

    q = Booking.query.filter(Booking.status != "cancelled")

    if search_date:
        try:
            ref_date = datetime.strptime(search_date, "%Y-%m-%d").date()
            q = q.filter(Booking.delivery_date == ref_date)
        except ValueError:
            pass

    if query_text:
        filt = booking_search_filter(query_text)
        if filt is not None:
            q = q.filter(filt)

    results = q.order_by(Booking.delivery_date.desc()).limit(100).all()

    data = []
    for b in results:
        items_list = booking_items_for_api(b)
        data.append({
            "id": b.id,
            "booking_number": b.booking_number,
            "serial": b.monthly_serial,
            "customer_name": b.customer_name,
            "contact_1": b.contact_1,
            "delivery_date": b.delivery_date.isoformat(),
            "delivery_time": b.delivery_time,
            "return_date": b.return_date.isoformat(),
            "status": b.status,
            "total_price": b.total_price,
            "venue": b.venue or "",
            "items": items_list,
        })
    return jsonify(data)


# ────────────────── BOOKING DELIVERY ──────────────────
@app.route("/booking-delivery")
@login_required
def booking_delivery_page():
    return render_template("delivery/delivery.html")


@app.route("/api/delivery/search")
def api_delivery_search():
    """Search bookings for delivery panel with smart month-based suggestion."""
    search_date = request.args.get("date", date.today().isoformat())
    query_text = request.args.get("q", "").strip()

    try:
        ref_date = datetime.strptime(search_date, "%Y-%m-%d").date()
    except ValueError:
        ref_date = date.today()

    year = ref_date.year
    month = ref_date.month

    def get_bookings_for_month(y, m):
        m_start = date(y, m, 1)
        if m == 12:
            m_end = date(y + 1, 1, 1)
        else:
            m_end = date(y, m + 1, 1)
        q = Booking.query.filter(
            Booking.delivery_date >= m_start,
            Booking.delivery_date < m_end,
            Booking.status.in_(["booked"]),
        )
        if query_text:
            filt = booking_search_filter(query_text)
            if filt is not None:
                q = q.filter(filt)
        return q.order_by(Booking.delivery_date).all()

    results = get_bookings_for_month(year, month)
    if not results and query_text:
        prev_m = month - 1 if month > 1 else 12
        prev_y = year if month > 1 else year - 1
        next_m = month + 1 if month < 12 else 1
        next_y = year if month < 12 else year + 1
        results = get_bookings_for_month(prev_y, prev_m) + get_bookings_for_month(next_y, next_m)

    data = []
    for b in results:
        items_list = booking_items_for_api(b)
        data.append({
            "id": b.id,
            "booking_number": b.booking_number,
            "serial": b.monthly_serial,
            "customer_name": b.customer_name,
            "contact_1": b.contact_1,
            "whatsapp_no": b.whatsapp_no,
            "delivery_date": b.delivery_date.isoformat(),
            "delivery_time": b.delivery_time,
            "return_date": b.return_date.isoformat(),
            "return_time": b.return_time,
            "status": b.status,
            "total_price": b.total_price,
            "total_advance": b.total_advance,
            "total_remaining": b.total_remaining,
            "venue": b.venue or "",
            "items": items_list,
        })
    return jsonify(data)


@app.route("/booking-delivery/<int:booking_id>")
@login_required
def booking_delivery_detail(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    # Check if any item has a next booking on return date (warning)
    next_bookings = []
    for bi in booking.booking_items:
        nxt = Booking.query.filter(
            Booking.id != booking.id,
            Booking.delivery_date == booking.return_date,
            Booking.status != "cancelled",
        ).join(BookingItem).filter(BookingItem.item_id == bi.item_id).first()
        if nxt:
            next_bookings.append({
                "dress": dress_display_name(bi.dress_name, bi.category, booking_item_size(bi)),
                "next_customer": nxt.customer_name,
                "next_serial": nxt.monthly_serial,
                "next_time": nxt.delivery_time,
                "next_venue": nxt.venue or "",
            })
    return render_template("delivery/detail.html", booking=booking, next_bookings=next_bookings)


@app.route("/booking-delivery/<int:booking_id>/save", methods=["POST"])
@login_required
def booking_delivery_save(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    remaining_collected = float(request.form.get("remaining_collected", 0) or 0)
    security_collected = float(request.form.get("security_collected", 0) or 0)
    delivery_notes = request.form.get("delivery_notes", "")
    mark_delivered = request.form.get("mark_delivered") == "1"

    booking.remaining_collected = remaining_collected
    booking.security_collected = security_collected
    booking.delivery_notes = delivery_notes
    if mark_delivered and booking.status == "booked":
        booking.status = "delivered"
        booking.delivered_at = datetime.now()

    db.session.commit()
    flash("Delivery details saved!", "success")
    return redirect(url_for("booking_delivery_detail", booking_id=booking.id))


# ────────────────── RETURN ──────────────────
@app.route("/return")
@login_required
def booking_return_page():
    return render_template("delivery/return.html")


@app.route("/api/return/search")
def api_return_search():
    """Search delivered bookings for return panel."""
    search_date = request.args.get("date", date.today().isoformat())
    query_text = request.args.get("q", "").strip()

    try:
        ref_date = datetime.strptime(search_date, "%Y-%m-%d").date()
    except ValueError:
        ref_date = date.today()

    year = ref_date.year
    month = ref_date.month

    def get_bookings_for_month(y, m):
        m_start = date(y, m, 1)
        if m == 12:
            m_end = date(y + 1, 1, 1)
        else:
            m_end = date(y, m + 1, 1)
        q = Booking.query.filter(
            Booking.return_date >= m_start,
            Booking.return_date < m_end,
            Booking.status == "delivered",
        )
        if query_text:
            filt = booking_search_filter(query_text)
            if filt is not None:
                q = q.filter(filt)
        return q.order_by(Booking.return_date).all()

    results = get_bookings_for_month(year, month)
    if not results and query_text:
        prev_m = month - 1 if month > 1 else 12
        prev_y = year if month > 1 else year - 1
        results = get_bookings_for_month(prev_y, prev_m)

    data = []
    for b in results:
        items_list = booking_items_for_api(b)
        data.append({
            "id": b.id,
            "booking_number": b.booking_number,
            "serial": b.monthly_serial,
            "customer_name": b.customer_name,
            "contact_1": b.contact_1,
            "whatsapp_no": b.whatsapp_no,
            "delivery_date": b.delivery_date.isoformat(),
            "return_date": b.return_date.isoformat(),
            "return_time": b.return_time,
            "status": b.status,
            "total_price": b.total_price,
            "total_remaining": b.total_remaining,
            "remaining_collected": b.remaining_collected or 0,
            "venue": b.venue or "",
            "items": items_list,
        })
    return jsonify(data)


@app.route("/return/<int:booking_id>")
@login_required
def booking_return_detail(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    return render_template("delivery/return_detail.html", booking=booking)


@app.route("/return/<int:booking_id>/save", methods=["POST"])
@login_required
def booking_return_save(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    action = request.form.get("action", "")

    if action == "mark_returned":
        booking.status = "returned"
        booking.returned_at = datetime.now()
        for bi in booking.booking_items:
            item = ClothingItem.query.get(bi.item_id)
            if item:
                item.status = "available"
        db.session.commit()
        flash("Booking marked as RETURNED. Items are now available.", "success")

    elif action == "incomplete_return":
        booking.status = "incomplete_return"
        booking.incomplete_notes = request.form.get("incomplete_notes", "")
        booking.security_held = float(request.form.get("security_held", 0) or 0)
        booking.returned_at = datetime.now()
        for bi in booking.booking_items:
            item = ClothingItem.query.get(bi.item_id)
            if item:
                item.status = "available"
        db.session.commit()
        flash("Marked as INCOMPLETE RETURN. Security held.", "warning")

    return redirect(url_for("booking_return_detail", booking_id=booking.id))


# ────────────────── INCOMPLETE RETURN ──────────────────
@app.route("/incomplete-return")
@login_required
def incomplete_return_page():
    bookings = Booking.query.filter_by(status="incomplete_return").order_by(Booking.returned_at.desc()).all()
    return render_template("delivery/incomplete_return.html", bookings=bookings)


@app.route("/incomplete-return/<int:booking_id>/resolve", methods=["POST"])
@login_required
def incomplete_return_resolve(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    if booking.status == "incomplete_return":
        booking.status = "returned"
        db.session.commit()
        flash(f"Booking Serial #{booking.monthly_serial:02d} resolved. Security ₹{booking.security_held:,.0f} deducted.", "success")
    return redirect(url_for("incomplete_return_page"))


# ────────────────── LATE RETURN ──────────────────
@app.route("/late-return")
@login_required
def late_return_page():
    today_val = date.today()
    late_bookings = Booking.query.filter(
        Booking.return_date < today_val,
        Booking.status == "delivered",
    ).order_by(Booking.return_date).all()
    return render_template("delivery/late_return.html", bookings=late_bookings, today=today_val)


# ────────────────── RETURNING TODAY (Other menu) ──────────────────
@app.route("/returning-today")
@login_required
def returning_today_page():
    return render_template("delivery/returning_today.html")


@app.route("/api/returning-today")
def api_returning_today():
    target_date = request.args.get("date", date.today().isoformat())
    try:
        ref_date = datetime.strptime(target_date, "%Y-%m-%d").date()
    except ValueError:
        ref_date = date.today()

    returning = Booking.query.filter(
        Booking.return_date == ref_date,
        Booking.status.in_(["booked", "delivered"]),
    ).order_by(Booking.return_time).all()

    def item_ids_of(bk):
        if bk.booking_items:
            return [bi.item_id for bi in bk.booking_items]
        return [bk.item_id] if bk.item_id else []

    def item_names_of(bk):
        if bk.booking_items:
            return [bi.dress_name for bi in bk.booking_items]
        return [bk.dress_name] if bk.dress_name else []

    records = []
    for b in returning:
        ret_item_ids = item_ids_of(b)
        next_booking_info = None
        # Find a booking delivering this same item on the same date
        candidates = Booking.query.filter(
            Booking.id != b.id,
            Booking.delivery_date == ref_date,
            Booking.status.in_(["booked", "delivered"]),
        ).all()
        for nxt in candidates:
            if set(item_ids_of(nxt)) & set(ret_item_ids):
                matched = list(set(item_ids_of(nxt)) & set(ret_item_ids))
                matched_names = [bi.dress_name for bi in nxt.booking_items if bi.item_id in matched] or item_names_of(nxt)
                next_booking_info = {
                    "customer_name": nxt.customer_name,
                    "contact_1": nxt.contact_1,
                    "whatsapp_no": nxt.whatsapp_no or "",
                    "address": nxt.customer_address,
                    "delivery_time": nxt.delivery_time,
                    "delivery_date": nxt.delivery_date.isoformat(),
                    "venue": nxt.venue or "",
                    "total_price": nxt.total_price,
                    "serial": nxt.monthly_serial,
                    "items": matched_names,
                }
                break

        records.append({
            "id": b.id,
            "serial": b.monthly_serial,
            "customer_name": b.customer_name,
            "contact_1": b.contact_1,
            "whatsapp_no": b.whatsapp_no,
            "address": b.customer_address,
            "return_date": b.return_date.isoformat(),
            "return_time": b.return_time,
            "venue": b.venue or "",
            "security_collected": b.security_collected or 0,
            "total_remaining": b.total_remaining,
            "remaining_collected": b.remaining_collected or 0,
            "delivery_notes": b.delivery_notes or "",
            "items": item_names_of(b),
            "next_booking": next_booking_info,
        })
    return jsonify(records)


# ────────────────── FINANCE: SECURITY DEPOSIT ──────────────────
@app.route("/finance/security-deposit")
@owner_required
def finance_security_deposit():
    total_collected = db.session.query(db.func.coalesce(db.func.sum(Booking.security_collected), 0)).filter(
        Booking.status.in_(["delivered", "returned", "incomplete_return"]),
    ).scalar()
    total_held = db.session.query(db.func.coalesce(db.func.sum(Booking.security_held), 0)).filter(
        Booking.status == "incomplete_return",
    ).scalar()
    total_returned = total_collected - total_held
    bookings = Booking.query.filter(
        Booking.security_collected > 0,
        Booking.status.in_(["delivered", "returned", "incomplete_return"]),
    ).order_by(Booking.delivered_at.desc()).all()
    return render_template("finance/security_deposit.html",
                           total_collected=total_collected,
                           total_held=total_held,
                           total_returned=total_returned,
                           bookings=bookings)


# ────────────────── BOOKED ITEMS: order by delivery date ──────────────────
# (Already handled in api_booking_list sort – ensure ordering)


# ─── Category Management ───────────────────────────────────────────────────────

@app.route("/manage-categories")
@login_required
def manage_categories():
    custom_cats = CustomCategory.query.order_by(CustomCategory.group, CustomCategory.name).all()
    base = {
        "mens": _BASE_MENS,
        "womens": _BASE_WOMENS,
        "jewellery": _BASE_JEWELLERY,
        "accessory": _BASE_ACCESSORY,
    }
    return render_template("manage_categories.html", custom_cats=custom_cats, base=base)


@app.route("/manage-categories/add", methods=["POST"])
@login_required
def add_category():
    name = request.form.get("name", "").strip()
    group = request.form.get("group", "other").strip()
    if not name:
        flash("Category name is required.", "error")
        return redirect(url_for("manage_categories"))
    existing = CustomCategory.query.filter_by(name=name).first()
    if existing:
        if not existing.active:
            existing.active = True
            db.session.commit()
            flash(f"Category '{name}' re-enabled.", "success")
        else:
            flash(f"Category '{name}' already exists.", "error")
        return redirect(url_for("manage_categories"))
    # Check it's not a base category
    all_base = _BASE_MENS + _BASE_WOMENS + _BASE_JEWELLERY + _BASE_ACCESSORY + ["Other"]
    if name in all_base:
        flash(f"'{name}' is already a built-in category.", "error")
        return redirect(url_for("manage_categories"))
    cat = CustomCategory(name=name, group=group)
    db.session.add(cat)
    db.session.commit()
    flash(f"Category '{name}' added to {group} group.", "success")
    return redirect(url_for("manage_categories"))


@app.route("/manage-categories/<int:id>/remove", methods=["POST"])
@owner_required
def remove_category(id):
    cat = CustomCategory.query.get_or_404(id)
    cat.active = False
    db.session.commit()
    flash(f"Category '{cat.name}' removed.", "success")
    return redirect(url_for("manage_categories"))


# ─── User Management (Owner only) ─────────────────────────────────────────────

@app.route("/users")
@owner_required
def users_page():
    users = User.query.order_by(User.role.desc(), User.username).all()
    linked_ids = [u.staff_id for u in users if u.staff_id]
    if linked_ids:
        staff_without_account = Staff.query.filter(Staff.active == True, ~Staff.id.in_(linked_ids)).all()
    else:
        staff_without_account = Staff.query.filter_by(active=True).all()
    return render_template("auth/users.html", users=users, staff_list=staff_without_account)


@app.route("/users/<int:id>/change-role", methods=["POST"])
@owner_required
def change_user_role(id):
    user = User.query.get_or_404(id)
    me = get_current_user()
    if user.id == me.id:
        flash("You cannot change your own role.", "error")
        return redirect(url_for("users_page"))
    new_role = request.form.get("role", "staff")
    user.role = new_role
    db.session.commit()
    flash(f"Role for '{user.username}' updated to '{new_role}'.", "success")
    return redirect(url_for("users_page"))


@app.route("/users/<int:id>/reset-password", methods=["POST"])
@owner_required
def reset_user_password(id):
    user = User.query.get_or_404(id)
    new_password = request.form.get("password", "").strip()
    if not new_password or len(new_password) < 4:
        flash("Password must be at least 4 characters.", "error")
        return redirect(url_for("users_page"))
    user.set_password(new_password)
    db.session.commit()
    flash(f"Password for '{user.username}' has been reset.", "success")
    return redirect(url_for("users_page"))


@app.route("/users/<int:id>/deactivate", methods=["POST"])
@owner_required
def deactivate_user(id):
    user = User.query.get_or_404(id)
    me = get_current_user()
    if user.id == me.id:
        flash("You cannot deactivate your own account.", "error")
        return redirect(url_for("users_page"))
    user.active = not user.active
    db.session.commit()
    state = "activated" if user.active else "deactivated"
    flash(f"User '{user.username}' {state}.", "success")
    return redirect(url_for("users_page"))


@app.route("/profile/change-password", methods=["GET", "POST"])
@login_required
def change_own_password():
    user = get_current_user()
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "").strip()
        confirm = request.form.get("confirm_password", "").strip()
        if not user.check_password(current):
            flash("Current password is incorrect.", "error")
        elif len(new_pw) < 4:
            flash("New password must be at least 4 characters.", "error")
        elif new_pw != confirm:
            flash("Passwords do not match.", "error")
        else:
            user.set_password(new_pw)
            db.session.commit()
            flash("Password changed successfully.", "success")
            return redirect(url_for("dashboard"))
    return render_template("auth/change_password.html")


def apply_migrations():
    """Auto-apply schema migrations for new columns."""
    import sqlite3
    db_path = app.config["SQLALCHEMY_DATABASE_URI"].replace("sqlite:///", "")
    if not os.path.exists(db_path):
        return
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    columns_to_add = [
        ("bookings", "delivery_notes", "TEXT"),
        ("bookings", "remaining_collected", "REAL DEFAULT 0"),
        ("bookings", "security_collected", "REAL DEFAULT 0"),
        ("bookings", "delivered_at", "DATETIME"),
        ("bookings", "returned_at", "DATETIME"),
        ("bookings", "incomplete_notes", "TEXT"),
        ("bookings", "security_held", "REAL DEFAULT 0"),
        ("booking_items", "prepared_by", "TEXT"),
        ("booking_items", "checked_by", "TEXT"),
        ("booking_items", "is_packed_ready", "INTEGER DEFAULT 0"),
        ("booking_items", "packing_note", "TEXT"),
        ("supplier_purchases", "gst_amount", "REAL DEFAULT 0"),
        ("supplier_purchases", "transaction_type", "TEXT DEFAULT 'purchase'"),
    ]
    for table, col, col_type in columns_to_add:
        try:
            cursor.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}")
        except sqlite3.OperationalError:
            pass
    conn.commit()
    conn.close()


if __name__ == "__main__":
    with app.app_context():
        apply_migrations()
        db.create_all()
        seed_database()
        ensure_owner_exists()
    app.run(debug=True, port=5000)
